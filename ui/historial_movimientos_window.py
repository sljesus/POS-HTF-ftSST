"""
Ventana de Historial de Movimientos de Inventario para HTF POS
Usando componentes reutilizables del sistema de diseño
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QLineEdit, QSizePolicy, QFrame,
    QComboBox, QDateEdit, QLabel
)
from PySide6.QtCore import Qt, Signal, QDate, QThread, QTimer
from PySide6.QtGui import QFont
from datetime import datetime, timedelta
import logging

# Importar componentes del sistema de diseño
from ui.components import (
    WindowsPhoneTheme,
    TileButton,
    create_page_layout,
    ContentPanel,
    StyledLabel,
    SearchBar,
    show_info_dialog,
    show_warning_dialog,
    show_error_dialog,
    aplicar_estilo_fecha
)


class MovimientosLoaderThread(QThread):
    """Hilo para cargar movimientos de forma asíncrona"""
    
    movimientos_loaded = Signal(list)
    error_occurred = Signal(str)
    
    def __init__(self, pg_manager):
        super().__init__()
        self.pg_manager = pg_manager
    
    def run(self):
        """Cargar movimientos desde la base de datos en un hilo separado"""
        try:
            # Usar el método de postgres_manager que retorna movimientos completos
            rows = self.pg_manager.obtener_movimientos_completos(limite=1000)
            self.movimientos_loaded.emit(rows)
                
        except Exception as e:
            self.error_occurred.emit(str(e))


class HistorialMovimientosWindow(QWidget):
    """Widget para ver el historial completo de movimientos de inventario"""
    
    cerrar_solicitado = Signal()
    
    def __init__(self, pg_manager, supabase_service, user_data, parent=None):
        super().__init__(parent)
        self.pg_manager = pg_manager
        self.supabase_service = supabase_service
        self.user_data = user_data
        self.movimientos_data = []
        self.movimientos_filtrados = []
        self.loader_thread = None
        
        self.setup_ui()
        self.cargar_movimientos()
    
    def setup_ui(self):
        """Configurar interfaz del historial"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Contenido
        content = QWidget()
        content_layout = create_page_layout("HISTORIAL DE MOVIMIENTOS")
        content.setLayout(content_layout)
        
        # Panel de filtros
        filters_panel = self.create_filters_panel()
        content_layout.addWidget(filters_panel)
        
        # Panel para la tabla
        table_panel = self.create_table_panel()
        content_layout.addWidget(table_panel)
        
        # Panel de información y botones
        info_buttons_panel = self.create_info_buttons_panel()
        content_layout.addWidget(info_buttons_panel)
        
        layout.addWidget(content)
    
    def create_filters_panel(self):
        """Crear el panel de filtros"""
        filters_panel = ContentPanel()
        filters_layout = QVBoxLayout(filters_panel)
        
        # Primera fila de filtros
        filters_row1 = QHBoxLayout()
        filters_row1.setSpacing(WindowsPhoneTheme.MARGIN_MEDIUM)
        
        # Buscador
        self.search_bar = SearchBar("Buscar por código, nombre o usuario...")
        self.search_bar.connect_search(self.aplicar_filtros)
        filters_row1.addWidget(self.search_bar, stretch=3)
        
        # Filtro por tipo de movimiento
        tipo_container = self.create_tipo_filter()
        filters_row1.addWidget(tipo_container, stretch=1)
        
        filters_layout.addLayout(filters_row1)
        
        # Segunda fila - Rango de fechas
        filters_row2 = self.create_fecha_filters()
        filters_layout.addLayout(filters_row2)
        
        return filters_panel
    
    def create_tipo_filter(self):
        """Crear el filtro por tipo de movimiento"""
        tipo_container = QWidget()
        tipo_layout = QVBoxLayout(tipo_container)
        tipo_layout.setContentsMargins(0, 0, 0, 0)
        tipo_layout.setSpacing(4)
        
        tipo_label = StyledLabel("Tipo:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        tipo_layout.addWidget(tipo_label)
        
        self.tipo_combo = QComboBox()
        self.tipo_combo.addItems([
            "Todos",
            "Entrada",
            "Salida",
            "Ajuste"
        ])
        self.tipo_combo.setMinimumHeight(40)
        self.tipo_combo.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.tipo_combo.currentTextChanged.connect(self.aplicar_filtros)
        tipo_layout.addWidget(self.tipo_combo)
        
        return tipo_container
    
    def create_fecha_filters(self):
        """Crear los filtros de fecha"""
        filters_row2 = QHBoxLayout()
        filters_row2.setSpacing(WindowsPhoneTheme.MARGIN_MEDIUM)
        
        # Fecha inicio
        fecha_inicio_container = QWidget()
        fecha_inicio_layout = QVBoxLayout(fecha_inicio_container)
        fecha_inicio_layout.setContentsMargins(0, 0, 0, 0)
        fecha_inicio_layout.setSpacing(4)
        
        fecha_inicio_label = StyledLabel("Desde:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        fecha_inicio_layout.addWidget(fecha_inicio_label)
        
        self.fecha_inicio = QDateEdit()
        self.fecha_inicio.setDate(QDate.currentDate().addMonths(-1))
        self.fecha_inicio.setCalendarPopup(True)
        aplicar_estilo_fecha(self.fecha_inicio)
        self.fecha_inicio.setMinimumHeight(40)
        self.fecha_inicio.dateChanged.connect(self.aplicar_filtros)
        fecha_inicio_layout.addWidget(self.fecha_inicio)
        
        filters_row2.addWidget(fecha_inicio_container, stretch=1)
        
        # Fecha fin
        fecha_fin_container = QWidget()
        fecha_fin_layout = QVBoxLayout(fecha_fin_container)
        fecha_fin_layout.setContentsMargins(0, 0, 0, 0)
        fecha_fin_layout.setSpacing(4)
        
        fecha_fin_label = StyledLabel("Hasta:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        fecha_fin_layout.addWidget(fecha_fin_label)
        
        self.fecha_fin = QDateEdit()
        self.fecha_fin.setDate(QDate.currentDate())
        self.fecha_fin.setCalendarPopup(True)
        aplicar_estilo_fecha(self.fecha_fin)
        self.fecha_fin.setMinimumHeight(40)
        self.fecha_fin.dateChanged.connect(self.aplicar_filtros)
        fecha_fin_layout.addWidget(self.fecha_fin)
        
        filters_row2.addWidget(fecha_fin_container, stretch=1)
        
        # Botón limpiar filtros
        btn_limpiar = QPushButton("Limpiar Filtros")
        btn_limpiar.setMinimumHeight(40)
        btn_limpiar.setMinimumWidth(120)
        btn_limpiar.setObjectName("tileButton")
        btn_limpiar.setProperty("tileColor", WindowsPhoneTheme.TILE_ORANGE)
        btn_limpiar.clicked.connect(self.limpiar_filtros)
        filters_row2.addWidget(btn_limpiar, alignment=Qt.AlignBottom)
        
        return filters_row2
    
    def create_table_panel(self):
        """Crear el panel con la tabla de movimientos"""
        table_panel = ContentPanel()
        table_layout = QVBoxLayout(table_panel)
        table_layout.setContentsMargins(0, 0, 0, 0)
        
        # Tabla de movimientos
        self.movimientos_table = QTableWidget()
        self.movimientos_table.setColumnCount(9)
        self.movimientos_table.setHorizontalHeaderLabels([
            "Fecha", "Tipo", "Código", "Producto", "Cantidad", 
            "Stock Ant.", "Stock Nuevo", "Motivo", "Usuario"
        ])
        
        # Configurar header
        header = self.movimientos_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.Stretch)
        header.setSectionResizeMode(8, QHeaderView.ResizeToContents)
        
        # Estilo de la tabla
        self.movimientos_table.setAlternatingRowColors(True)
        self.movimientos_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.movimientos_table.setSelectionMode(QTableWidget.SingleSelection)
        self.movimientos_table.verticalHeader().setVisible(False)
        self.movimientos_table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        table_layout.addWidget(self.movimientos_table)
        return table_panel
    
    def create_info_buttons_panel(self):
        """Crear el panel de información y botones"""
        info_buttons_panel = ContentPanel()
        info_buttons_layout = QHBoxLayout(info_buttons_panel)
        
        # Etiqueta de información
        self.info_label = StyledLabel("", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        info_buttons_layout.addWidget(self.info_label, stretch=1)
        
        # Botones de acción
        btn_exportar = TileButton("Exportar Excel", "fa5s.file-excel", WindowsPhoneTheme.TILE_GREEN)
        btn_exportar.clicked.connect(self.exportar_excel)
        btn_exportar.setMaximumWidth(200)
        info_buttons_layout.addWidget(btn_exportar)
        
        btn_actualizar = TileButton("Actualizar", "fa5s.sync", WindowsPhoneTheme.TILE_BLUE)
        btn_actualizar.clicked.connect(self.cargar_movimientos)
        btn_actualizar.setMaximumWidth(200)
        info_buttons_layout.addWidget(btn_actualizar)
        
        btn_volver = TileButton("Volver", "fa5s.arrow-left", WindowsPhoneTheme.TILE_RED)
        btn_volver.clicked.connect(self.cerrar_solicitado.emit)
        btn_volver.setMaximumWidth(200)
        info_buttons_layout.addWidget(btn_volver)
        
        return info_buttons_panel
    
    def cargar_movimientos(self):
        """Cargar todos los movimientos desde la base de datos de forma asíncrona"""
        try:
            # Mostrar indicador de carga
            self.info_label.setText("Cargando movimientos...")
            self.movimientos_table.setRowCount(0)
            
            # Detener hilo anterior si existe
            if self.loader_thread and self.loader_thread.isRunning():
                self.loader_thread.quit()
                self.loader_thread.wait(1000)  # Esperar hasta 1 segundo
            
            # Crear y ejecutar hilo de carga
            self.loader_thread = MovimientosLoaderThread(self.pg_manager)
            self.loader_thread.movimientos_loaded.connect(self.procesar_datos_movimientos)
            self.loader_thread.error_occurred.connect(self.mostrar_error_carga)
            self.loader_thread.finished.connect(self.on_thread_finished)
            self.loader_thread.start()
            
        except Exception as e:
            logging.error(f"Error iniciando carga de movimientos: {e}")
            show_error_dialog(
                self,
                "Error al cargar",
                "No se pudieron cargar los movimientos",
                detail=str(e)
            )
    
    def on_thread_finished(self):
        """Callback cuando el thread termina"""
        pass
    
    def procesar_datos_movimientos(self, rows):
        """Procesar los datos de movimientos cargados desde la base de datos"""
        try:
            self.movimientos_data = []
            
            for row in rows:
                self.movimientos_data.append({
                    'id_movimiento': row['id_movimiento'],
                    'fecha': row['fecha'],
                    'tipo_movimiento': row['tipo_movimiento'],
                    'codigo_interno': row['codigo_interno'],
                    'tipo_producto': row['tipo_producto'],
                    'cantidad': row['cantidad'],
                    'stock_anterior': row['stock_anterior'],
                    'stock_nuevo': row['stock_nuevo'],
                    'motivo': row['motivo'] or '',
                    'id_usuario': row['id_usuario'],
                    'id_venta': row['id_venta'],
                    'nombre_producto': row['nombre_producto'],
                    'nombre_usuario': row['nombre_usuario'] or 'Usuario desconocido'
                })
            
            self.aplicar_filtros()
            logging.info(f"Movimientos cargados: {len(self.movimientos_data)}")
            
        except Exception as e:
            logging.error(f"Error procesando datos de movimientos: {e}")
            show_error_dialog(
                self,
                "Error al procesar datos",
                "No se pudieron procesar los datos de movimientos",
                detail=str(e)
            )
    
    def mostrar_error_carga(self, error_msg):
        """Mostrar mensaje de error al cargar movimientos"""
        logging.error(f"Error cargando movimientos: {error_msg}")
        show_error_dialog(
            self,
            "Error al cargar",
            "No se pudieron cargar los movimientos",
            detail=error_msg
        )
        self.info_label.setText("Error al cargar movimientos")
    
    def aplicar_filtros(self):
        """Aplicar todos los filtros activos"""
        try:
            # Obtener criterios de filtro
            texto_busqueda = self.search_bar.text().strip().lower()
            tipo_seleccionado = self.tipo_combo.currentText()
            fecha_desde = self.fecha_inicio.date().toPython()
            fecha_hasta = self.fecha_fin.date().toPython()
            
            # Filtrar datos
            self.movimientos_filtrados = []
            for mov in self.movimientos_data:
                # Filtro de texto
                if texto_busqueda:
                    if not any([
                        texto_busqueda in mov['codigo_interno'].lower(),
                        texto_busqueda in mov['nombre_producto'].lower(),
                        texto_busqueda in mov['nombre_usuario'].lower(),
                        texto_busqueda in (mov['motivo'] or '').lower()
                    ]):
                        continue
                
                # Filtro de tipo
                if tipo_seleccionado != "Todos":
                    if mov['tipo_movimiento'].lower() != tipo_seleccionado.lower():
                        continue
                
                # Filtro de fecha
                fecha_mov = mov['fecha'].date() if isinstance(mov['fecha'], datetime) else mov['fecha']
                if not (fecha_desde <= fecha_mov <= fecha_hasta):
                    continue
                
                self.movimientos_filtrados.append(mov)
            
            self.mostrar_movimientos(self.movimientos_filtrados)
            
        except Exception as e:
            logging.error(f"Error aplicando filtros: {e}")
            self.mostrar_movimientos(self.movimientos_data)
    
    def mostrar_movimientos(self, movimientos):
        """Mostrar movimientos en la tabla"""
        try:
            self.movimientos_table.setRowCount(0)
            
            for mov in movimientos:
                row = self.movimientos_table.rowCount()
                self.movimientos_table.insertRow(row)
                
                # Fecha
                fecha_str = mov['fecha'].strftime("%d/%m/%Y %H:%M") if isinstance(mov['fecha'], datetime) else str(mov['fecha'])
                item_fecha = QTableWidgetItem(fecha_str)
                item_fecha.setTextAlignment(Qt.AlignCenter)
                self.movimientos_table.setItem(row, 0, item_fecha)
                
                # Tipo de movimiento
                tipo = mov['tipo_movimiento'].capitalize()
                item_tipo = QTableWidgetItem(tipo)
                item_tipo.setTextAlignment(Qt.AlignCenter)
                
                # Color según tipo
                if mov['tipo_movimiento'].lower() == 'entrada':
                    item_tipo.setForeground(Qt.darkGreen)
                elif mov['tipo_movimiento'].lower() == 'salida':
                    item_tipo.setForeground(Qt.darkRed)
                else:
                    item_tipo.setForeground(Qt.darkBlue)
                
                self.movimientos_table.setItem(row, 1, item_tipo)
                
                # Código interno
                self.movimientos_table.setItem(row, 2, QTableWidgetItem(mov['codigo_interno']))
                
                # Nombre producto
                self.movimientos_table.setItem(row, 3, QTableWidgetItem(mov['nombre_producto']))
                
                # Cantidad
                cantidad = mov['cantidad']
                item_cantidad = QTableWidgetItem(str(cantidad))
                item_cantidad.setTextAlignment(Qt.AlignCenter)
                if cantidad > 0:
                    item_cantidad.setForeground(Qt.darkGreen)
                else:
                    item_cantidad.setForeground(Qt.darkRed)
                self.movimientos_table.setItem(row, 4, item_cantidad)
                
                # Stock anterior
                item_stock_ant = QTableWidgetItem(str(mov['stock_anterior']))
                item_stock_ant.setTextAlignment(Qt.AlignCenter)
                self.movimientos_table.setItem(row, 5, item_stock_ant)
                
                # Stock nuevo
                item_stock_nuevo = QTableWidgetItem(str(mov['stock_nuevo']))
                item_stock_nuevo.setTextAlignment(Qt.AlignCenter)
                self.movimientos_table.setItem(row, 6, item_stock_nuevo)
                
                # Motivo
                self.movimientos_table.setItem(row, 7, QTableWidgetItem(mov['motivo']))
                
                # Usuario
                item_usuario = QTableWidgetItem(mov['nombre_usuario'])
                item_usuario.setTextAlignment(Qt.AlignCenter)
                self.movimientos_table.setItem(row, 8, item_usuario)
            
            # Actualizar información
            total_movimientos = len(movimientos)
            total_general = len(self.movimientos_data)
            
            if total_movimientos == total_general:
                self.info_label.setText(f"Total de movimientos: {total_movimientos}")
            else:
                self.info_label.setText(f"Mostrando {total_movimientos} de {total_general} movimientos")
            
            logging.info(f"Mostrando {total_movimientos} movimientos en tabla")
            
        except Exception as e:
            logging.error(f"Error mostrando movimientos: {e}")
            show_error_dialog(
                self,
                "Error de visualización",
                "No se pudieron mostrar los movimientos",
                detail=str(e)
            )
    
    def limpiar_filtros(self):
        """Limpiar todos los filtros y mostrar todo"""
        self.search_bar.clear()
        self.tipo_combo.setCurrentIndex(0)
        self.fecha_inicio.setDate(QDate.currentDate().addMonths(-1))
        self.fecha_fin.setDate(QDate.currentDate())
        self.aplicar_filtros()
    
    def exportar_excel(self):
        """Exportar movimientos filtrados a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Movimientos Inventario"
            
            # Encabezados
            headers = [
                "Fecha", "Tipo", "Código", "Producto", "Cantidad", 
                "Stock Anterior", "Stock Nuevo", "Motivo", "Usuario", "ID Venta"
            ]
            
            # Estilo de encabezado
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Obtener movimientos filtrados actuales
            texto_busqueda = self.search_bar.text().strip().lower()
            tipo_seleccionado = self.tipo_combo.currentText()
            fecha_desde = self.fecha_inicio.date().toPython()
            fecha_hasta = self.fecha_fin.date().toPython()
            
            movimientos_exportar = []
            for mov in self.movimientos_data:
                if texto_busqueda:
                    if not any([
                        texto_busqueda in mov['codigo_interno'].lower(),
                        texto_busqueda in mov['nombre_producto'].lower(),
                        texto_busqueda in mov['nombre_usuario'].lower(),
                        texto_busqueda in (mov['motivo'] or '').lower()
                    ]):
                        continue
                
                if tipo_seleccionado != "Todos":
                    if mov['tipo_movimiento'].lower() != tipo_seleccionado.lower():
                        continue
                
                fecha_mov = mov['fecha'].date() if isinstance(mov['fecha'], datetime) else mov['fecha']
                if not (fecha_desde <= fecha_mov <= fecha_hasta):
                    continue
                
                movimientos_exportar.append(mov)
            
            # Datos
            for row_idx, mov in enumerate(movimientos_exportar, start=2):
                fecha_str = mov['fecha'].strftime("%d/%m/%Y %H:%M") if isinstance(mov['fecha'], datetime) else str(mov['fecha'])
                
                ws.cell(row=row_idx, column=1, value=fecha_str)
                ws.cell(row=row_idx, column=2, value=mov['tipo_movimiento'].capitalize())
                ws.cell(row=row_idx, column=3, value=mov['codigo_interno'])
                ws.cell(row=row_idx, column=4, value=mov['nombre_producto'])
                ws.cell(row=row_idx, column=5, value=mov['cantidad'])
                ws.cell(row=row_idx, column=6, value=mov['stock_anterior'])
                ws.cell(row=row_idx, column=7, value=mov['stock_nuevo'])
                ws.cell(row=row_idx, column=8, value=mov['motivo'])
                ws.cell(row=row_idx, column=9, value=mov['nombre_usuario'])
                ws.cell(row=row_idx, column=10, value=mov['id_venta'] if mov['id_venta'] else '')
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 30
            ws.column_dimensions['I'].width = 20
            ws.column_dimensions['J'].width = 10
            
            # Guardar archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"movimientos_inventario_{timestamp}.xlsx"
            wb.save(filename)
            
            show_info_dialog(
                self,
                "Exportación exitosa",
                f"Archivo generado: {filename}\n\nMovimientos exportados: {len(movimientos_exportar)}"
            )
            
            logging.info(f"Reporte de movimientos exportado: {filename}")
            
        except ImportError:
            show_error_dialog(
                self,
                "Módulo no disponible",
                "No se puede exportar a Excel. El módulo openpyxl no está instalado."
            )
        except Exception as e:
            logging.error(f"Error exportando a Excel: {e}")
            show_error_dialog(
                self,
                "Error de exportación",
                "No se pudo generar el archivo Excel",
                detail=str(e)
            )
    
    def closeEvent(self, event):
        """Evento al cerrar la ventana"""
        # Detener hilo de carga si está activo
        if self.loader_thread and self.loader_thread.isRunning():
            self.loader_thread.quit()
            if not self.loader_thread.wait(2000):  # Esperar hasta 2 segundos
                self.loader_thread.terminate()
                self.loader_thread.wait()
            
        super().closeEvent(event)
    
    def __del__(self):
        """Destructor para limpiar el thread"""
        try:
            if hasattr(self, 'loader_thread') and self.loader_thread:
                if self.loader_thread.isRunning():
                    self.loader_thread.quit()
                    self.loader_thread.wait(1000)
        except (RuntimeError, AttributeError):
            # El thread ya fue eliminado por C++
            pass