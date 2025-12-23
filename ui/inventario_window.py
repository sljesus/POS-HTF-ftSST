"""
Ventana de Grid de Inventario para HTF POS
Usando componentes reutilizables del sistema de diseño
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QLineEdit, QSizePolicy, QFrame,
    QComboBox, QCheckBox, QDialog, QAbstractItemView
)
from PySide6.QtCore import Qt, Signal, QTimer
from PySide6.QtGui import QFont
import logging

# Importar componentes del sistema de diseño
from ui.components import (
    WindowsPhoneTheme,
    TileButton,
    create_page_layout,
    ContentPanel,
    StyledLabel,
    SearchBar,
    show_info_dialog,
    show_warning_dialog,
    show_error_dialog
)
from ui.editable_catalog_grid import EditableCatalogGrid


class InventarioWindow(QWidget):
    """Widget para ver el grid completo de inventario"""
    
    cerrar_solicitado = Signal()
    
    def __init__(self, postgres_manager, supabase_service, user_data, parent=None):
        super().__init__(parent)
        self.pg_manager = postgres_manager  # Cambiado de db_manager a pg_manager
        self.supabase_service = supabase_service
        self.user_data = user_data
        self.productos_data = []
        
        # Timer para detectar entrada del escáner
        self.scanner_timer = QTimer()
        self.scanner_timer.setSingleShot(True)
        self.scanner_timer.setInterval(300)  # 300ms después de que deje de escribir
        self.scanner_timer.timeout.connect(self.filtrar_inventario)
        
        self.setup_ui()
        self.cargar_inventario()
    
    def setup_ui(self):
        """Configurar interfaz de inventario"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Contenido
        content = QWidget()
        content_layout = create_page_layout("INVENTARIO COMPLETO")
        content.setLayout(content_layout)
        
        # Buscador
        self.search_bar = SearchBar("Buscar por código interno, código de barras, nombre o categoría...")
        self.search_bar.connect_search(self.on_search_changed)
        content_layout.addWidget(self.search_bar)
        
        # Panel de filtros
        filters_panel = ContentPanel()
        filters_layout = QHBoxLayout(filters_panel)
        filters_layout.setSpacing(WindowsPhoneTheme.MARGIN_MEDIUM)
        
        # Filtro por categoría
        categoria_container = QWidget()
        categoria_layout = QVBoxLayout(categoria_container)
        categoria_layout.setContentsMargins(0, 0, 0, 0)
        categoria_layout.setSpacing(4)
        
        categoria_label = StyledLabel("Categoría:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        categoria_layout.addWidget(categoria_label)
        
        self.categoria_combo = QComboBox()
        self.categoria_combo.setMinimumHeight(40)
        self.categoria_combo.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.categoria_combo.currentTextChanged.connect(self.aplicar_filtros)
        categoria_layout.addWidget(self.categoria_combo)
        
        filters_layout.addWidget(categoria_container, stretch=1)
        
        # Filtro por tipo de producto
        tipo_container = QWidget()
        tipo_layout = QVBoxLayout(tipo_container)
        tipo_layout.setContentsMargins(0, 0, 0, 0)
        tipo_layout.setSpacing(4)
        
        tipo_label = StyledLabel("Tipo:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        tipo_layout.addWidget(tipo_label)
        
        self.tipo_combo = QComboBox()
        self.tipo_combo.addItems(["Todos", "Producto Varios", "Suplemento"])
        self.tipo_combo.setMinimumHeight(40)
        self.tipo_combo.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.tipo_combo.currentTextChanged.connect(self.aplicar_filtros)
        tipo_layout.addWidget(self.tipo_combo)
        
        filters_layout.addWidget(tipo_container, stretch=1)
        
        # Filtro por estado de stock
        stock_container = QWidget()
        stock_layout = QVBoxLayout(stock_container)
        stock_layout.setContentsMargins(0, 0, 0, 0)
        stock_layout.setSpacing(4)
        
        stock_label = StyledLabel("Stock:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        stock_layout.addWidget(stock_label)
        
        self.stock_combo = QComboBox()
        self.stock_combo.addItems(["Todos", "Bajo Stock", "Sin Stock", "Stock Normal"])
        self.stock_combo.setMinimumHeight(40)
        self.stock_combo.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.stock_combo.currentTextChanged.connect(self.aplicar_filtros)
        stock_layout.addWidget(self.stock_combo)
        
        filters_layout.addWidget(stock_container, stretch=1)
        
        # Filtro por ubicación
        ubicacion_container = QWidget()
        ubicacion_layout = QVBoxLayout(ubicacion_container)
        ubicacion_layout.setContentsMargins(0, 0, 0, 0)
        ubicacion_layout.setSpacing(4)
        
        ubicacion_label = StyledLabel("Ubicación:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        ubicacion_layout.addWidget(ubicacion_label)
        
        self.ubicacion_combo = QComboBox()
        self.ubicacion_combo.setMinimumHeight(40)
        self.ubicacion_combo.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.ubicacion_combo.currentTextChanged.connect(self.aplicar_filtros)
        ubicacion_layout.addWidget(self.ubicacion_combo)
        
        filters_layout.addWidget(ubicacion_container, stretch=1)
        
        # Checkbox solo activos
        activos_container = QWidget()
        activos_layout = QVBoxLayout(activos_container)
        activos_layout.setContentsMargins(0, 0, 0, 0)
        activos_layout.setSpacing(4)
        
        activos_spacer = StyledLabel("", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        activos_layout.addWidget(activos_spacer)
        
        self.check_solo_activos = QCheckBox("Solo activos")
        self.check_solo_activos.setChecked(True)
        self.check_solo_activos.setMinimumHeight(40)
        self.check_solo_activos.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.check_solo_activos.setStyleSheet(f"""
            QCheckBox {{
                spacing: 8px;
                color: #333;
            }}
            QCheckBox::indicator {{
                width: 20px;
                height: 20px;
                border: 2px solid #cccccc;
                border-radius: 3px;
                background-color: white;
            }}
            QCheckBox::indicator:checked {{
                background-color: {WindowsPhoneTheme.TILE_BLUE};
                border-color: {WindowsPhoneTheme.TILE_BLUE};
                image: none;
            }}
            QCheckBox::indicator:checked:after {{
                content: "✓";
                color: white;
                font-weight: bold;
            }}
        """)
        self.check_solo_activos.stateChanged.connect(self.aplicar_filtros)
        activos_layout.addWidget(self.check_solo_activos)
        
        filters_layout.addWidget(activos_container, stretch=1)
        
        # Botón limpiar filtros
        btn_limpiar_container = QWidget()
        btn_limpiar_layout = QVBoxLayout(btn_limpiar_container)
        btn_limpiar_layout.setContentsMargins(0, 0, 0, 0)
        btn_limpiar_layout.setSpacing(4)
        
        limpiar_spacer = StyledLabel("", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        btn_limpiar_layout.addWidget(limpiar_spacer)
        
        btn_limpiar = QPushButton("Limpiar")
        btn_limpiar.setMinimumHeight(40)
        btn_limpiar.setMinimumWidth(100)
        btn_limpiar.setObjectName("tileButton")
        btn_limpiar.setProperty("tileColor", WindowsPhoneTheme.TILE_ORANGE)
        btn_limpiar.clicked.connect(self.limpiar_filtros)
        btn_limpiar_layout.addWidget(btn_limpiar)
        
        filters_layout.addWidget(btn_limpiar_container)
        
        content_layout.addWidget(filters_panel)
        
        # Panel para la tabla
        table_panel = ContentPanel()
        table_layout = QVBoxLayout(table_panel)
        table_layout.setContentsMargins(0, 0, 0, 0)
        
        # Tabla de inventario
        self.inventory_table = QTableWidget()
        self.inventory_table.setColumnCount(8)
        self.inventory_table.setHorizontalHeaderLabels([
            "Código", "Nombre", "Categoría", "Precio", "Stock", "Stock Min", "Ubicación", "Estado"
        ])
        
        # Configurar header
        header = self.inventory_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)
        
        self.inventory_table.verticalHeader().setVisible(False)
        self.inventory_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.inventory_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.inventory_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.inventory_table.setAlternatingRowColors(True)
        self.inventory_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # Aplicar estilos a la tabla
        self.inventory_table.setStyleSheet(f"""
            QTableWidget {{
                background-color: white;
                border: none;
                gridline-color: #e5e7eb;
            }}
            QTableWidget::item {{
                padding: 8px;
                border-bottom: 1px solid #e5e7eb;
            }}
            QTableWidget::item:selected {{
                background-color: {WindowsPhoneTheme.TILE_BLUE};
                color: white;
            }}
            QHeaderView::section {{
                background-color: {WindowsPhoneTheme.BG_LIGHT};
                color: {WindowsPhoneTheme.PRIMARY_BLUE};
                padding: 12px 8px;
                border: none;
                border-bottom: 2px solid {WindowsPhoneTheme.TILE_BLUE};
                font-weight: bold;
                font-family: '{WindowsPhoneTheme.FONT_FAMILY}';
                font-size: {WindowsPhoneTheme.FONT_SIZE_NORMAL}px;
            }}
        """)
        
        table_layout.addWidget(self.inventory_table)
        content_layout.addWidget(table_panel)
        
        # Panel de información
        info_panel = ContentPanel()
        info_layout = QHBoxLayout(info_panel)
        
        self.info_label = StyledLabel("", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        info_layout.addWidget(self.info_label, stretch=1)
        
        content_layout.addWidget(info_panel)
        
        # Botones de acción
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(WindowsPhoneTheme.TILE_SPACING)
        
        btn_editar_catalogo = TileButton("Editar Catálogo", "fa5s.edit", WindowsPhoneTheme.TILE_PURPLE)
        btn_editar_catalogo.clicked.connect(self.abrir_grid_editable)
        
        btn_reporte = TileButton("Generar Reporte", "fa5s.file-excel", WindowsPhoneTheme.TILE_GREEN)
        btn_reporte.clicked.connect(self.generar_reporte)
        
        btn_bajo_stock = TileButton("Ver Bajo Stock", "fa5s.exclamation-triangle", WindowsPhoneTheme.TILE_ORANGE)
        btn_bajo_stock.clicked.connect(self.filtrar_bajo_stock)
        
        btn_cerrar = TileButton("Cerrar", "fa5s.times", WindowsPhoneTheme.TILE_RED)
        btn_cerrar.clicked.connect(self.cerrar_solicitado.emit)
        
        buttons_layout.addWidget(btn_editar_catalogo)
        buttons_layout.addWidget(btn_reporte)
        buttons_layout.addWidget(btn_bajo_stock)
        buttons_layout.addWidget(btn_cerrar)
        
        content_layout.addLayout(buttons_layout)
        layout.addWidget(content)
    
    def cargar_inventario(self):
        """Cargar datos de inventario desde la base de datos"""
        try:
            logging.info("Cargando inventario completo...")
            # Usar el método de postgres_manager en lugar de acceso directo
            self.productos_data = self.pg_manager.obtener_inventario_completo()
            
            # Poblar combo de categorías
            categorias = sorted(set(p.get('categoria') for p in self.productos_data if p.get('categoria')))
            self.categoria_combo.clear()
            self.categoria_combo.addItem("Todas")
            self.categoria_combo.addItems(categorias)
            
            # Poblar combo de ubicaciones
            ubicaciones = sorted(set(p.get('ubicacion') for p in self.productos_data if p.get('ubicacion')))
            self.ubicacion_combo.clear()
            self.ubicacion_combo.addItem("Todas")
            self.ubicacion_combo.addItems(ubicaciones)
            
            self.aplicar_filtros()
            
            logging.info(f"Inventario cargado: {len(self.productos_data)} productos")
            
        except Exception as e:
            logging.error(f"Error cargando inventario: {e}")
            show_error_dialog(
                self,
                "Error al cargar",
                "No se pudo cargar el inventario",
                detail=str(e)
            )
    
    def mostrar_inventario(self, productos):
        """Mostrar productos en la tabla"""
        self.inventory_table.setRowCount(0)
        
        for row, producto in enumerate(productos):
            self.inventory_table.insertRow(row)
            
            # Código
            self.inventory_table.setItem(row, 0, QTableWidgetItem(producto['codigo_interno']))
            
            # Nombre
            self.inventory_table.setItem(row, 1, QTableWidgetItem(producto['nombre']))
            
            # Categoría (usar seccion si está disponible)
            self.inventory_table.setItem(row, 2, QTableWidgetItem(producto.get('seccion', 'N/A')))
            
            # Precio
            precio = float(producto['precio']) if producto['precio'] is not None else 0.0
            precio_item = QTableWidgetItem(f"${precio:.2f}")
            precio_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.inventory_table.setItem(row, 3, precio_item)
            
            # Stock actual
            stock_item = QTableWidgetItem(str(producto['stock_actual']))
            stock_item.setTextAlignment(Qt.AlignCenter)
            
            # Colorear si está bajo en stock
            if producto['stock_actual'] <= producto['stock_minimo']:
                stock_item.setForeground(Qt.red)
            
            self.inventory_table.setItem(row, 4, stock_item)
            
            # Stock mínimo
            stock_min_item = QTableWidgetItem(str(producto['stock_minimo']))
            stock_min_item.setTextAlignment(Qt.AlignCenter)
            self.inventory_table.setItem(row, 5, stock_min_item)
            
            # Ubicación
            ubicacion = producto.get('ubicacion', 'N/A')
            self.inventory_table.setItem(row, 6, QTableWidgetItem(ubicacion))
            
            # Estado
            estado = "Activo" if producto['activo'] else "Inactivo"
            estado_item = QTableWidgetItem(estado)
            if not producto['activo']:
                estado_item.setForeground(Qt.gray)
            self.inventory_table.setItem(row, 7, estado_item)
        
        # Actualizar información
        total_productos = len(productos)
        total_general = len(self.productos_data)
        
        if total_productos == total_general:
            self.info_label.setText(f"Total de productos: {total_productos}")
        else:
            self.info_label.setText(f"Mostrando {total_productos} de {total_general} productos")
    
    def on_search_changed(self):
        """Manejar cambio en el campo de búsqueda con timer para escáner"""
        # Reiniciar el timer cada vez que se escribe
        self.scanner_timer.stop()
        self.scanner_timer.start()
    
    def aplicar_filtros(self):
        """Aplicar todos los filtros seleccionados"""
        try:
            texto_busqueda = self.search_bar.text().lower()
            categoria_seleccionada = self.categoria_combo.currentText()
            tipo_seleccionado = self.tipo_combo.currentText()
            stock_seleccionado = self.stock_combo.currentText()
            ubicacion_seleccionada = self.ubicacion_combo.currentText()
            solo_activos = self.check_solo_activos.isChecked()
            
            productos_filtrados = []
            
            for producto in self.productos_data:
                # Filtro de búsqueda de texto
                if texto_busqueda:
                    texto_match = (
                        texto_busqueda in producto['codigo_interno'].lower()
                        or texto_busqueda in producto['nombre'].lower()
                        or (producto.get('seccion') and texto_busqueda in producto['seccion'].lower())
                        or (producto.get('codigo_barras') and texto_busqueda in producto['codigo_barras'].lower())
                    )
                    if not texto_match:
                        continue
                
                # Filtro de categoría (usar seccion)
                if categoria_seleccionada != "Todas":
                    if producto.get('seccion') != categoria_seleccionada:
                        continue
                
                # Filtro de tipo de producto
                if tipo_seleccionado == "Producto Varios":
                    if producto['tipo_producto'] != 'varios':
                        continue
                elif tipo_seleccionado == "Suplemento":
                    if producto['tipo_producto'] != 'suplemento':
                        continue
                
                # Filtro de estado de stock
                if stock_seleccionado == "Bajo Stock":
                    if producto['stock_actual'] > producto['stock_minimo']:
                        continue
                elif stock_seleccionado == "Sin Stock":
                    if producto['stock_actual'] > 0:
                        continue
                elif stock_seleccionado == "Stock Normal":
                    if producto['stock_actual'] <= producto['stock_minimo']:
                        continue
                
                # Filtro de ubicación
                if ubicacion_seleccionada != "Todas":
                    if producto.get('ubicacion') != ubicacion_seleccionada:
                        continue
                
                # Filtro de activos
                if solo_activos and not producto['activo']:
                    continue
                
                productos_filtrados.append(producto)
            
            self.mostrar_inventario(productos_filtrados)
            
        except Exception as e:
            logging.error(f"Error aplicando filtros: {e}")
            self.mostrar_inventario(self.productos_data)
    
    def filtrar_inventario(self):
        """Filtrar inventario (llamado por el timer del escáner)"""
        self.aplicar_filtros()
    
    def limpiar_filtros(self):
        """Limpiar todos los filtros"""
        self.search_bar.clear()
        self.categoria_combo.setCurrentIndex(0)
        self.tipo_combo.setCurrentIndex(0)
        self.stock_combo.setCurrentIndex(0)
        self.ubicacion_combo.setCurrentIndex(0)
        self.check_solo_activos.setChecked(True)
        self.aplicar_filtros()
    
    def filtrar_bajo_stock(self):
        """Filtrar productos con stock bajo o menor al mínimo"""
        productos_bajo_stock = [
            p for p in self.productos_data
            if p['stock_actual'] <= p['stock_minimo']
        ]
        
        if productos_bajo_stock:
            self.mostrar_inventario(productos_bajo_stock)
            show_info_dialog(
                self,
                "Productos bajo stock",
                f"Se encontraron {len(productos_bajo_stock)} productos con stock bajo o menor al mínimo"
            )
        else:
            show_info_dialog(
                self,
                "Stock saludable",
                "No hay productos con stock bajo en este momento"
            )
    
    def generar_reporte(self):
        """Generar reporte de inventario en Excel"""
        try:
            from datetime import datetime
            import os
            
            # Verificar si openpyxl está disponible
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                show_warning_dialog(
                    self,
                    "Biblioteca requerida",
                    "Para generar reportes en Excel necesitas instalar openpyxl",
                    detail="Ejecuta: pip install openpyxl"
                )
                return
            
            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Inventario"
            
            # Estilos
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezados
            headers = ["Código", "Nombre", "Categoría", "Precio", "Stock Actual", "Stock Mín", "Ubicación", "Estado"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Datos
            for row, producto in enumerate(self.productos_data, 2):
                ws.cell(row=row, column=1, value=producto['codigo_interno']).border = border
                ws.cell(row=row, column=2, value=producto['nombre']).border = border
                ws.cell(row=row, column=3, value=producto.get('seccion', 'N/A')).border = border
                
                precio_cell = ws.cell(row=row, column=4, value=producto['precio'])
                precio_cell.number_format = '$#,##0.00'
                precio_cell.border = border
                
                stock_cell = ws.cell(row=row, column=5, value=producto['stock_actual'])
                stock_cell.alignment = Alignment(horizontal="center")
                stock_cell.border = border
                
                # Resaltar bajo stock
                if producto['stock_actual'] <= producto['stock_minimo']:
                    stock_cell.font = Font(color="FF0000", bold=True)
                
                stock_min_cell = ws.cell(row=row, column=6, value=producto['stock_minimo'])
                stock_min_cell.alignment = Alignment(horizontal="center")
                stock_min_cell.border = border
                
                ws.cell(row=row, column=7, value=producto.get('ubicacion', 'N/A')).border = border
                ws.cell(row=row, column=8, value="Activo" if producto['activo'] else "Inactivo").border = border
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 12
            
            # Guardar archivo
            fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            filename = os.path.join(desktop, f"Inventario_HTF_{fecha_str}.xlsx")
            
            wb.save(filename)
            
            show_info_dialog(
                self,
                "Reporte generado",
                f"El reporte de inventario ha sido generado exitosamente",
                detail=f"Archivo guardado en:\n{filename}"
            )
            
            logging.info(f"Reporte de inventario generado: {filename}")
            
        except Exception as e:
            logging.error(f"Error generando reporte: {e}")
            show_error_dialog(
                self,
                "Error al generar reporte",
                "No se pudo crear el archivo de Excel",
                detail=str(e)
            )
    
    def abrir_grid_editable(self):
        """Abrir ventana con grid editable del catálogo"""
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("Editar Catálogo de Productos")
            dialog.setMinimumSize(1200, 700)
            
            layout = QVBoxLayout(dialog)
            layout.setContentsMargins(0, 0, 0, 0)
            
            # Crear grid editable
            grid_editable = EditableCatalogGrid(self.pg_manager)
            grid_editable.catalogo_actualizado.connect(self.cargar_inventario)
            
            layout.addWidget(grid_editable)
            
            dialog.exec()
            
        except Exception as e:
            logging.error(f"Error abriendo grid editable: {e}")
            show_error_dialog(
                self,
                "Error",
                "No se pudo abrir el editor de catálogo",
                detail=str(e)
            )