"""
Ventana de Historial de Acceso al Gimnasio para HTF POS
Usando componentes reutilizables del sistema de diseño
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QLineEdit, QSizePolicy, QFrame,
    QComboBox, QDateEdit, QLabel
)
from PySide6.QtCore import Qt, Signal, QDate, QThread, QTimer  # Eliminado pyqtSignal
from PySide6.QtGui import QFont
from datetime import datetime, timedelta
import logging

# Importar componentes del sistema de diseño
from ui.components import (
    WindowsPhoneTheme,
    TileButton,
    create_page_layout,
    ContentPanel,
    StyledLabel,
    SearchBar,
    show_info_dialog,
    show_warning_dialog,
    show_error_dialog,
    aplicar_estilo_fecha
)


class AccesosLoaderThread(QThread):
    """Hilo para cargar accesos de forma asíncrona"""
    
    accesos_loaded = Signal(list)  # Cambiado de pyqtSignal a Signal
    error_occurred = Signal(str)   # Cambiado de pyqtSignal a Signal
    
    def __init__(self, supabase_service):
        super().__init__()
        self.supabase_service = supabase_service
    
    def run(self):
        """Cargar accesos desde Supabase en un hilo separado"""
        try:
            if self.supabase_service and self.supabase_service.is_connected:
                # Consultar registro_entradas desde Supabase con relaciones
                response = self.supabase_service.client.table('registro_entradas')\
                    .select('''
                        *,
                        miembros(
                            nombres,
                            apellido_paterno,
                            apellido_materno,
                            codigo_qr
                        ),
                        personal(
                            nombres,
                            apellido_paterno,
                            apellido_materno,
                            id_personal,
                            numero_empleado
                        )
                    ''')\
                    .order('fecha_entrada', desc=True)\
                    .execute()
                
                rows = response.data if response.data else []
                self.accesos_loaded.emit(rows)
            else:
                self.error_occurred.emit("No hay conexión a Supabase")
        except Exception as e:
            self.error_occurred.emit(str(e))


class HistorialAccesoWindow(QWidget):
    """Widget para ver el historial completo de accesos al gimnasio"""
    
    cerrar_solicitado = Signal()
    
    def __init__(self, db_manager, supabase_service, user_data, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.supabase_service = supabase_service
        self.user_data = user_data
        self.accesos_data = []
        self.accesos_filtrados = []
        self.loader_thread = None
        self.update_timer = None
        
        self.setup_ui()
        self.cargar_accesos()
        
        # Configurar timer para actualizar tiempos en tiempo real
        self.setup_update_timer()
    
    def setup_ui(self):
        """Configurar interfaz del historial"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Contenido
        content = QWidget()
        content_layout = create_page_layout("HISTORIAL DE ACCESO")
        content.setLayout(content_layout)
        
        # Panel de filtros
        filters_panel = self.create_filters_panel()
        content_layout.addWidget(filters_panel)
        
        # Panel para la tabla
        table_panel = self.create_table_panel()
        content_layout.addWidget(table_panel)
        
        # Panel de información y botones
        info_buttons_panel = self.create_info_buttons_panel()
        content_layout.addWidget(info_buttons_panel)
        
        layout.addWidget(content)
    
    def create_filters_panel(self):
        """Crear el panel de filtros"""
        filters_panel = ContentPanel()
        filters_layout = QVBoxLayout(filters_panel)
        
        # Primera fila de filtros
        filters_row1 = QHBoxLayout()
        filters_row1.setSpacing(WindowsPhoneTheme.MARGIN_MEDIUM)
        
        # Buscador
        self.search_bar = SearchBar("Buscar por nombre, código o área...")
        self.search_bar.connect_search(self.aplicar_filtros)
        filters_row1.addWidget(self.search_bar, stretch=3)
        
        # Filtro por tipo de acceso
        tipo_container = self.create_tipo_filter()
        filters_row1.addWidget(tipo_container, stretch=1)
        
        # Filtro por estado (dentro/fuera)
        estado_container = self.create_estado_filter()
        filters_row1.addWidget(estado_container, stretch=1)
        
        filters_layout.addLayout(filters_row1)
        
        # Segunda fila - Rango de fechas
        filters_row2 = self.create_fecha_filters()
        filters_layout.addLayout(filters_row2)
        
        return filters_panel
    
    def create_tipo_filter(self):
        """Crear el filtro por tipo de acceso"""
        tipo_container = QWidget()
        tipo_layout = QVBoxLayout(tipo_container)
        tipo_layout.setContentsMargins(0, 0, 0, 0)
        tipo_layout.setSpacing(4)
        
        tipo_label = StyledLabel("Tipo:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        tipo_layout.addWidget(tipo_label)
        
        self.tipo_combo = QComboBox()
        self.tipo_combo.addItems([
            "Todos",
            "Miembro",
            "Personal",
            "Visitante"
        ])
        self.tipo_combo.setMinimumHeight(40)
        self.tipo_combo.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.tipo_combo.currentTextChanged.connect(self.aplicar_filtros)
        tipo_layout.addWidget(self.tipo_combo)
        
        return tipo_container
    
    def create_estado_filter(self):
        """Crear el filtro por estado"""
        estado_container = QWidget()
        estado_layout = QVBoxLayout(estado_container)
        estado_layout.setContentsMargins(0, 0, 0, 0)
        estado_layout.setSpacing(4)
        
        estado_label = StyledLabel("Estado:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        estado_layout.addWidget(estado_label)
        
        self.estado_combo = QComboBox()
        self.estado_combo.addItems([
            "Todos",
            "Dentro",
            "Salió"
        ])
        self.estado_combo.setMinimumHeight(40)
        self.estado_combo.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.estado_combo.currentTextChanged.connect(self.aplicar_filtros)
        estado_layout.addWidget(self.estado_combo)
        
        return estado_container
    
    def create_fecha_filters(self):
        """Crear los filtros de fecha"""
        filters_row2 = QHBoxLayout()
        filters_row2.setSpacing(WindowsPhoneTheme.MARGIN_MEDIUM)
        
        # Fecha inicio
        fecha_inicio_container = QWidget()
        fecha_inicio_layout = QVBoxLayout(fecha_inicio_container)
        fecha_inicio_layout.setContentsMargins(0, 0, 0, 0)
        fecha_inicio_layout.setSpacing(4)
        
        fecha_inicio_label = StyledLabel("Desde:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        fecha_inicio_layout.addWidget(fecha_inicio_label)
        
        self.fecha_inicio = QDateEdit()
        self.fecha_inicio.setDate(QDate.currentDate().addDays(-7))  # Últimos 7 días
        self.fecha_inicio.setCalendarPopup(True)
        aplicar_estilo_fecha(self.fecha_inicio)
        self.fecha_inicio.setMinimumHeight(40)
        self.fecha_inicio.dateChanged.connect(self.aplicar_filtros)
        fecha_inicio_layout.addWidget(self.fecha_inicio)
        
        filters_row2.addWidget(fecha_inicio_container, stretch=1)
        
        # Fecha fin
        fecha_fin_container = QWidget()
        fecha_fin_layout = QVBoxLayout(fecha_fin_container)
        fecha_fin_layout.setContentsMargins(0, 0, 0, 0)
        fecha_fin_layout.setSpacing(4)
        
        fecha_fin_label = StyledLabel("Hasta:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        fecha_fin_layout.addWidget(fecha_fin_label)
        
        self.fecha_fin = QDateEdit()
        self.fecha_fin.setDate(QDate.currentDate())
        self.fecha_fin.setCalendarPopup(True)
        aplicar_estilo_fecha(self.fecha_fin)
        self.fecha_fin.setMinimumHeight(40)
        self.fecha_fin.dateChanged.connect(self.aplicar_filtros)
        fecha_fin_layout.addWidget(self.fecha_fin)
        
        filters_row2.addWidget(fecha_fin_container, stretch=1)
        
        # Botón limpiar filtros
        btn_limpiar = QPushButton("Limpiar Filtros")
        btn_limpiar.setMinimumHeight(40)
        btn_limpiar.setMinimumWidth(120)
        btn_limpiar.setObjectName("tileButton")
        btn_limpiar.setProperty("tileColor", WindowsPhoneTheme.TILE_ORANGE)
        btn_limpiar.clicked.connect(self.limpiar_filtros)
        filters_row2.addWidget(btn_limpiar, alignment=Qt.AlignBottom)
        
        return filters_row2
    
    def create_table_panel(self):
        """Crear el panel con la tabla de accesos"""
        table_panel = ContentPanel()
        table_layout = QVBoxLayout(table_panel)
        table_layout.setContentsMargins(0, 0, 0, 0)
        
        # Tabla de accesos
        self.accesos_table = QTableWidget()
        self.accesos_table.setColumnCount(9)
        self.accesos_table.setHorizontalHeaderLabels([
            "Fecha Entrada", "Fecha Salida", "Tipo", "Nombre", 
            "Código", "Área", "Tiempo", "Dispositivo", "Notas"
        ])
        
        # Configurar header
        header = self.accesos_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(8, QHeaderView.Stretch)
        
        # Estilo de la tabla
        self.accesos_table.setAlternatingRowColors(True)
        self.accesos_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.accesos_table.setSelectionMode(QTableWidget.SingleSelection)
        self.accesos_table.verticalHeader().setVisible(False)
        
        table_layout.addWidget(self.accesos_table)
        return table_panel
    
    def create_info_buttons_panel(self):
        """Crear el panel de información y botones"""
        info_buttons_panel = ContentPanel()
        info_buttons_layout = QHBoxLayout(info_buttons_panel)
        
        # Etiqueta de información
        self.info_label = StyledLabel("", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        info_buttons_layout.addWidget(self.info_label, stretch=1)
        
        # Botones de acción
        btn_exportar = TileButton("Exportar Excel", "fa5s.file-excel", WindowsPhoneTheme.TILE_GREEN)
        btn_exportar.clicked.connect(self.exportar_excel)
        btn_exportar.setMaximumWidth(200)
        info_buttons_layout.addWidget(btn_exportar)
        
        btn_actualizar = TileButton("Actualizar", "fa5s.sync", WindowsPhoneTheme.TILE_BLUE)
        btn_actualizar.clicked.connect(self.cargar_accesos)
        btn_actualizar.setMaximumWidth(200)
        info_buttons_layout.addWidget(btn_actualizar)
        
        btn_volver = TileButton("Volver", "fa5s.arrow-left", WindowsPhoneTheme.TILE_RED)
        btn_volver.clicked.connect(self.cerrar_solicitado.emit)
        btn_volver.setMaximumWidth(200)
        info_buttons_layout.addWidget(btn_volver)
        
        return info_buttons_panel
    
    def setup_update_timer(self):
        """Configurar el timer para actualizar tiempos en tiempo real"""
        self.update_timer = QTimer(self)
        self.update_timer.timeout.connect(self.actualizar_tiempos)
        self.update_timer.start(60000)  # Actualizar cada minuto
    
    def cargar_accesos(self):
        """Cargar todos los accesos desde Supabase de forma asíncrona"""
        try:
            # Mostrar indicador de carga
            self.info_label.setText("Cargando accesos...")
            self.accesos_table.setRowCount(0)
            
            # Detener hilo anterior si existe
            if self.loader_thread and self.loader_thread.isRunning():
                self.loader_thread.terminate()
                self.loader_thread.wait()
            
            # Crear y ejecutar hilo de carga
            self.loader_thread = AccesosLoaderThread(self.supabase_service)
            self.loader_thread.accesos_loaded.connect(self.procesar_datos_accesos)
            self.loader_thread.error_occurred.connect(self.mostrar_error_carga)
            self.loader_thread.start()
            
        except Exception as e:
            logging.error(f"Error iniciando carga de accesos: {e}")
            show_error_dialog(
                self,
                "Error al cargar",
                "No se pudieron cargar los accesos",
                detail=str(e)
            )
    
    def procesar_datos_accesos(self, rows):
        """Procesar los datos de accesos cargados desde Supabase"""
        try:
            self.accesos_data = []
            
            for row in rows:
                # Determinar nombre completo según tipo de acceso
                tipo_acceso = row.get('tipo_acceso', '')
                nombre_completo = 'Desconocido'
                codigo = 'N/A'
                
                if tipo_acceso == 'miembro' and row.get('miembros'):
                    miembro = row['miembros']
                    nombre_completo = f"{miembro.get('nombres', '')} {miembro.get('apellido_paterno', '')} {miembro.get('apellido_materno', '')}".strip()
                    codigo = miembro.get('codigo_qr', 'N/A')
                elif tipo_acceso == 'personal' and row.get('personal'):
                    personal = row['personal']
                    nombre_completo = f"{personal.get('nombres', '')} {personal.get('apellido_paterno', '')} {personal.get('apellido_materno', '')}".strip()
                    codigo = personal.get('numero_empleado') or str(personal.get('id_personal', 'N/A'))
                elif tipo_acceso == 'visitante':
                    nombre_completo = row.get('nombre_visitante', 'Visitante')
                    codigo = 'N/A'
                
                # Procesar fechas
                fecha_entrada = row.get('fecha_entrada')
                if isinstance(fecha_entrada, str):
                    fecha_entrada = datetime.fromisoformat(fecha_entrada.replace('Z', '+00:00'))
                
                fecha_salida = row.get('fecha_salida')
                if isinstance(fecha_salida, str):
                    fecha_salida = datetime.fromisoformat(fecha_salida.replace('Z', '+00:00'))
                
                self.accesos_data.append({
                    'id_entrada': row.get('id_entrada'),
                    'fecha_entrada': fecha_entrada,
                    'fecha_salida': fecha_salida,
                    'tipo_acceso': tipo_acceso,
                    'area_accedida': row.get('area_accedida') or 'General',
                    'dispositivo_registro': row.get('dispositivo_registro') or 'Manual',
                    'notas': row.get('notas') or '',
                    'nombre_completo': nombre_completo,
                    'codigo': codigo
                })
            
            self.aplicar_filtros()
            logging.info(f"Accesos cargados: {len(self.accesos_data)}")
            
        except Exception as e:
            logging.error(f"Error procesando datos de accesos: {e}")
            show_error_dialog(
                self,
                "Error al procesar datos",
                "No se pudieron procesar los datos de accesos",
                detail=str(e)
            )
    
    def mostrar_error_carga(self, error_msg):
        """Mostrar mensaje de error al cargar accesos"""
        logging.error(f"Error cargando accesos: {error_msg}")
        show_error_dialog(
            self,
            "Error al cargar",
            "No se pudieron cargar los accesos",
            detail=error_msg
        )
        self.info_label.setText("Error al cargar accesos")
    
    def aplicar_filtros(self):
        """Aplicar todos los filtros activos"""
        try:
            # Obtener criterios de filtro
            texto_busqueda = self.search_bar.text().strip().lower()
            tipo_seleccionado = self.tipo_combo.currentText()
            estado_seleccionado = self.estado_combo.currentText()
            fecha_desde = self.fecha_inicio.date().toPython()
            fecha_hasta = self.fecha_fin.date().toPython()
            
            # Filtrar datos
            self.accesos_filtrados = []
            for acceso in self.accesos_data:
                # Filtro de texto
                if texto_busqueda:
                    if not any([
                        texto_busqueda in acceso['nombre_completo'].lower(),
                        texto_busqueda in acceso['codigo'].lower(),
                        texto_busqueda in acceso['area_accedida'].lower(),
                        texto_busqueda in (acceso['notas'] or '').lower()
                    ]):
                        continue
                
                # Filtro de tipo
                if tipo_seleccionado != "Todos":
                    if acceso['tipo_acceso'].lower() != tipo_seleccionado.lower():
                        continue
                
                # Filtro de estado (dentro/salió)
                if estado_seleccionado != "Todos":
                    if estado_seleccionado == "Dentro" and acceso['fecha_salida'] is not None:
                        continue
                    if estado_seleccionado == "Salió" and acceso['fecha_salida'] is None:
                        continue
                
                # Filtro de fecha
                fecha_acc = acceso['fecha_entrada'].date() if isinstance(acceso['fecha_entrada'], datetime) else acceso['fecha_entrada']
                if not (fecha_desde <= fecha_acc <= fecha_hasta):
                    continue
                
                self.accesos_filtrados.append(acceso)
            
            self.mostrar_accesos(self.accesos_filtrados)
            
        except Exception as e:
            logging.error(f"Error aplicando filtros: {e}")
            self.mostrar_accesos(self.accesos_data)
    
    def mostrar_accesos(self, accesos):
        """Mostrar accesos en la tabla"""
        try:
            self.accesos_table.setRowCount(0)
            
            for acceso in accesos:
                row = self.accesos_table.rowCount()
                self.accesos_table.insertRow(row)
                
                # Fecha entrada
                fecha_entrada_str = acceso['fecha_entrada'].strftime("%d/%m/%Y %H:%M") if isinstance(acceso['fecha_entrada'], datetime) else str(acceso['fecha_entrada'])
                item_entrada = QTableWidgetItem(fecha_entrada_str)
                item_entrada.setTextAlignment(Qt.AlignCenter)
                self.accesos_table.setItem(row, 0, item_entrada)
                
                # Fecha salida
                if acceso['fecha_salida']:
                    fecha_salida_str = acceso['fecha_salida'].strftime("%d/%m/%Y %H:%M") if isinstance(acceso['fecha_salida'], datetime) else str(acceso['fecha_salida'])
                    item_salida = QTableWidgetItem(fecha_salida_str)
                    item_salida.setTextAlignment(Qt.AlignCenter)
                else:
                    item_salida = QTableWidgetItem("DENTRO")
                    item_salida.setTextAlignment(Qt.AlignCenter)
                    item_salida.setForeground(Qt.darkGreen)
                    item_salida.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL, QFont.Bold))
                
                self.accesos_table.setItem(row, 1, item_salida)
                
                # Tipo de acceso
                tipo = acceso['tipo_acceso'].capitalize()
                item_tipo = QTableWidgetItem(tipo)
                item_tipo.setTextAlignment(Qt.AlignCenter)
                
                # Color según tipo
                if acceso['tipo_acceso'].lower() == 'miembro':
                    item_tipo.setForeground(Qt.darkBlue)
                elif acceso['tipo_acceso'].lower() == 'personal':
                    item_tipo.setForeground(Qt.darkGreen)
                else:
                    item_tipo.setForeground(Qt.darkRed)
                
                self.accesos_table.setItem(row, 2, item_tipo)
                
                # Nombre completo
                self.accesos_table.setItem(row, 3, QTableWidgetItem(acceso['nombre_completo']))
                
                # Código
                item_codigo = QTableWidgetItem(acceso['codigo'])
                item_codigo.setTextAlignment(Qt.AlignCenter)
                self.accesos_table.setItem(row, 4, item_codigo)
                
                # Área
                item_area = QTableWidgetItem(acceso['area_accedida'])
                item_area.setTextAlignment(Qt.AlignCenter)
                self.accesos_table.setItem(row, 5, item_area)
                
                # Tiempo de permanencia
                item_tiempo = QTableWidgetItem("-")
                item_tiempo.setTextAlignment(Qt.AlignCenter)
                self.accesos_table.setItem(row, 6, item_tiempo)
                
                # Dispositivo
                item_dispositivo = QTableWidgetItem(acceso['dispositivo_registro'])
                item_dispositivo.setTextAlignment(Qt.AlignCenter)
                self.accesos_table.setItem(row, 7, item_dispositivo)
                
                # Notas
                self.accesos_table.setItem(row, 8, QTableWidgetItem(acceso['notas']))
            
            # Actualizar información
            total_accesos = len(accesos)
            total_general = len(self.accesos_data)
            dentro_ahora = sum(1 for a in accesos if a['fecha_salida'] is None)
            
            if total_accesos == total_general:
                self.info_label.setText(f"Total de accesos: {total_accesos} | Dentro ahora: {dentro_ahora}")
            else:
                self.info_label.setText(f"Mostrando {total_accesos} de {total_general} accesos | Dentro ahora: {dentro_ahora}")
            
            logging.info(f"Mostrando {total_accesos} accesos en tabla")
            
        except Exception as e:
            logging.error(f"Error mostrando accesos: {e}")
            show_error_dialog(
                self,
                "Error de visualización",
                "No se pudieron mostrar los accesos",
                detail=str(e)
            )
    
    def actualizar_tiempos(self):
        """Actualizar los tiempos de permanencia para quienes aún están dentro"""
        try:
            # Solo actualizar si hay accesos mostrados
            if self.accesos_table.rowCount() == 0:
                return
                
            ahora = datetime.now()
            
            # Recorrer todas las filas de la tabla
            for row in range(self.accesos_table.rowCount()):
                # Obtener el item de la columna de fecha de salida
                item_salida = self.accesos_table.item(row, 1)
                
                # Si está "DENTRO", actualizar el tiempo
                if item_salida and item_salida.text() == "DENTRO":
                    # Obtener la fecha de entrada
                    item_entrada = self.accesos_table.item(row, 0)
                    if item_entrada:
                        try:
                            # Extraer la fecha de entrada del texto
                            fecha_entrada_str = item_entrada.text()
                            fecha_entrada = datetime.strptime(fecha_entrada_str, "%d/%m/%Y %H:%M")
                            
                            # Calcular tiempo transcurrido
                            delta = ahora - fecha_entrada
                            horas = delta.seconds // 3600
                            minutos = (delta.seconds % 3600) // 60
                            
                            # Si ha pasado más de un día, mostrar días también
                            if delta.days > 0:
                                tiempo_str = f"{delta.days}d {horas}h {minutos}m"
                            else:
                                tiempo_str = f"{horas}h {minutos}m"
                            
                            # Actualizar el item de tiempo
                            item_tiempo = self.accesos_table.item(row, 6)
                            if item_tiempo:
                                item_tiempo.setText(tiempo_str)
                        except Exception as e:
                            logging.warning(f"Error actualizando tiempo en fila {row}: {e}")
                            continue
            
        except Exception as e:
            logging.error(f"Error actualizando tiempos: {e}")
    
    def limpiar_filtros(self):
        """Limpiar todos los filtros y mostrar todo"""
        self.search_bar.clear()
        self.tipo_combo.setCurrentIndex(0)
        self.estado_combo.setCurrentIndex(0)
        self.fecha_inicio.setDate(QDate.currentDate().addDays(-7))
        self.fecha_fin.setDate(QDate.currentDate())
        self.aplicar_filtros()
    
    def exportar_excel(self):
        """Exportar accesos filtrados a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Historial Acceso"
            
            # Encabezados
            headers = [
                "Fecha Entrada", "Fecha Salida", "Tipo", "Nombre", 
                "Código", "Área", "Tiempo", "Dispositivo", "Notas"
            ]
            
            # Estilo de encabezado
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Obtener accesos filtrados actuales
            texto_busqueda = self.search_bar.text().strip().lower()
            tipo_seleccionado = self.tipo_combo.currentText()
            estado_seleccionado = self.estado_combo.currentText()
            fecha_desde = self.fecha_inicio.date().toPython()
            fecha_hasta = self.fecha_fin.date().toPython()
            
            accesos_exportar = []
            for acceso in self.accesos_data:
                if texto_busqueda:
                    if not any([
                        texto_busqueda in acceso['nombre_completo'].lower(),
                        texto_busqueda in acceso['codigo'].lower(),
                        texto_busqueda in acceso['area_accedida'].lower(),
                        texto_busqueda in (acceso['notas'] or '').lower()
                    ]):
                        continue
                
                if tipo_seleccionado != "Todos":
                    if acceso['tipo_acceso'].lower() != tipo_seleccionado.lower():
                        continue
                
                if estado_seleccionado != "Todos":
                    if estado_seleccionado == "Dentro" and acceso['fecha_salida'] is not None:
                        continue
                    if estado_seleccionado == "Salió" and acceso['fecha_salida'] is None:
                        continue
                
                fecha_acc = acceso['fecha_entrada'].date() if isinstance(acceso['fecha_entrada'], datetime) else acceso['fecha_entrada']
                if not (fecha_desde <= fecha_acc <= fecha_hasta):
                    continue
                
                accesos_exportar.append(acceso)
            
            # Datos
            for row_idx, acceso in enumerate(accesos_exportar, start=2):
                fecha_entrada_str = acceso['fecha_entrada'].strftime("%d/%m/%Y %H:%M") if isinstance(acceso['fecha_entrada'], datetime) else str(acceso['fecha_entrada'])
                fecha_salida_str = acceso['fecha_salida'].strftime("%d/%m/%Y %H:%M") if acceso['fecha_salida'] and isinstance(acceso['fecha_salida'], datetime) else ("DENTRO" if not acceso['fecha_salida'] else str(acceso['fecha_salida']))
                
                # Calcular tiempo
                if acceso['fecha_salida']:
                    try:
                        delta = acceso['fecha_salida'] - acceso['fecha_entrada']
                        horas = delta.seconds // 3600
                        minutos = (delta.seconds % 3600) // 60
                        tiempo_str = f"{horas}h {minutos}m"
                    except:
                        tiempo_str = "-"
                else:
                    tiempo_str = "-"
                
                ws.cell(row=row_idx, column=1, value=fecha_entrada_str)
                ws.cell(row=row_idx, column=2, value=fecha_salida_str)
                ws.cell(row=row_idx, column=3, value=acceso['tipo_acceso'].capitalize())
                ws.cell(row=row_idx, column=4, value=acceso['nombre_completo'])
                ws.cell(row=row_idx, column=5, value=acceso['codigo'])
                ws.cell(row=row_idx, column=6, value=acceso['area_accedida'])
                ws.cell(row=row_idx, column=7, value=tiempo_str)
                ws.cell(row=row_idx, column=8, value=acceso['dispositivo_registro'])
                ws.cell(row=row_idx, column=9, value=acceso['notas'])
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 30
            
            # Guardar archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"historial_acceso_{timestamp}.xlsx"
            wb.save(filename)
            
            show_info_dialog(
                self,
                "Exportación exitosa",
                f"Archivo generado: {filename}\n\nAccesos exportados: {len(accesos_exportar)}"
            )
            
            logging.info(f"Reporte de accesos exportado: {filename}")
            
        except ImportError:
            show_error_dialog(
                self,
                "Módulo no disponible",
                "No se puede exportar a Excel. El módulo openpyxl no está instalado."
            )
        except Exception as e:
            logging.error(f"Error exportando a Excel: {e}")
            show_error_dialog(
                self,
                "Error de exportación",
                "No se pudo generar el archivo Excel",
                detail=str(e)
            )
    
    def closeEvent(self, event):
        """Evento al cerrar la ventana"""
        # Detener hilo de carga si está activo
        if self.loader_thread and self.loader_thread.isRunning():
            self.loader_thread.terminate()
            self.loader_thread.wait()
        
        # Detener timer de actualización
        if self.update_timer:
            self.update_timer.stop()
            
        super().closeEvent(event)