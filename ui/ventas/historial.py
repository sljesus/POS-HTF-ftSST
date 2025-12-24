"""
Ventana de Historial de Ventas para HTF POS
Usando componentes reutilizables del sistema de diseño
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, 
    QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QDateEdit, QSizePolicy, QComboBox, QAbstractItemView
)
from PySide6.QtCore import Qt, Signal, QDate, QTimer
from PySide6.QtGui import QFont
import logging
from datetime import datetime
import qtawesome as qta

# Importar componentes del sistema de diseño
from ui.components import (
    WindowsPhoneTheme,
    TileButton,
    create_page_layout,
    ContentPanel,
    StyledLabel,
    SearchBar,
    show_info_dialog,
    show_warning_dialog,
    aplicar_estilo_fecha
)


class HistorialVentasWindow(QWidget):
    """Widget para ver historial de ventas"""
    
    cerrar_solicitado = Signal()
    
    def __init__(self, pg_manager, supabase_service, user_data, parent=None):
        super().__init__(parent)
        self.pg_manager = pg_manager
        self.supabase_service = supabase_service
        self.user_data = user_data
        self.ventas_data = []  # Almacenar todas las ventas cargadas
        
        # Timer para detectar entrada del escáner
        self.scanner_timer = QTimer()
        self.scanner_timer.setSingleShot(True)
        self.scanner_timer.setInterval(300)  # 300ms después de que deje de escribir
        self.scanner_timer.timeout.connect(self.aplicar_filtros)
        
        # Configurar política de tamaño
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        self.setup_ui()
        
    def setup_ui(self):
        """Configurar interfaz de historial"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Contenido
        content = QWidget()
        content_layout = create_page_layout("HISTORIAL COMPLETO")
        content.setLayout(content_layout)
        
        # Buscador
        self.search_bar = SearchBar("Buscar por ID de venta, usuario o monto...")
        self.search_bar.connect_search(self.on_search_changed)
        content_layout.addWidget(self.search_bar)
        
        # Filtros
        self.create_filters(content_layout)
        
        # Tabla
        self.create_history_table(content_layout)
        
        # Panel de información
        info_panel = ContentPanel()
        info_layout = QHBoxLayout(info_panel)
        self.info_label = StyledLabel("", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        info_layout.addWidget(self.info_label, stretch=1)
        content_layout.addWidget(info_panel)
        
        # Botones
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(WindowsPhoneTheme.TILE_SPACING)
        
        btn_exportar = TileButton("Exportar", "fa5s.download", WindowsPhoneTheme.TILE_GREEN)
        btn_exportar.clicked.connect(self.exportar_datos)
        
        btn_cerrar = TileButton("Cerrar", "fa5s.times", WindowsPhoneTheme.TILE_RED)
        btn_cerrar.clicked.connect(self.cerrar_solicitado.emit)
        
        buttons_layout.addWidget(btn_exportar)
        buttons_layout.addWidget(btn_cerrar)
        
        content_layout.addLayout(buttons_layout)
        layout.addWidget(content)
        
        # Cargar datos iniciales
        self.cargar_historial_completo()
        
    def create_filters(self, parent_layout):
        """Crear filtros de búsqueda"""
        filters_panel = ContentPanel()
        filters_layout = QHBoxLayout(filters_panel)
        filters_layout.setSpacing(WindowsPhoneTheme.MARGIN_MEDIUM)
        
        # Fecha desde
        desde_container = QWidget()
        desde_layout = QVBoxLayout(desde_container)
        desde_layout.setContentsMargins(0, 0, 0, 0)
        desde_layout.setSpacing(4)
        desde_label = StyledLabel("Desde:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        desde_layout.addWidget(desde_label)
        self.fecha_desde = QDateEdit()
        self.fecha_desde.setDate(QDate.currentDate().addDays(-30))
        self.fecha_desde.setCalendarPopup(True)
        self.fecha_desde.setMinimumHeight(40)
        self.fecha_desde.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.fecha_desde.dateChanged.connect(self.cargar_historial_completo)
        aplicar_estilo_fecha(self.fecha_desde)
        desde_layout.addWidget(self.fecha_desde)
        filters_layout.addWidget(desde_container, stretch=1)
        
        # Fecha hasta
        hasta_container = QWidget()
        hasta_layout = QVBoxLayout(hasta_container)
        hasta_layout.setContentsMargins(0, 0, 0, 0)
        hasta_layout.setSpacing(4)
        hasta_label = StyledLabel("Hasta:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        hasta_layout.addWidget(hasta_label)
        self.fecha_hasta = QDateEdit()
        self.fecha_hasta.setDate(QDate.currentDate())
        self.fecha_hasta.setCalendarPopup(True)
        self.fecha_hasta.setMinimumHeight(40)
        self.fecha_hasta.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.fecha_hasta.dateChanged.connect(self.cargar_historial_completo)
        aplicar_estilo_fecha(self.fecha_hasta)
        hasta_layout.addWidget(self.fecha_hasta)
        filters_layout.addWidget(hasta_container, stretch=1)
        
        # Filtro por usuario
        usuario_container = QWidget()
        usuario_layout = QVBoxLayout(usuario_container)
        usuario_layout.setContentsMargins(0, 0, 0, 0)
        usuario_layout.setSpacing(4)
        usuario_label = StyledLabel("Usuario:", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        usuario_layout.addWidget(usuario_label)
        self.usuario_combo = QComboBox()
        self.usuario_combo.setMinimumHeight(40)
        self.usuario_combo.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        self.usuario_combo.currentTextChanged.connect(self.aplicar_filtros)
        usuario_layout.addWidget(self.usuario_combo)
        filters_layout.addWidget(usuario_container, stretch=1)
        
        # Botón limpiar filtros
        btn_limpiar_container = QWidget()
        btn_limpiar_layout = QVBoxLayout(btn_limpiar_container)
        btn_limpiar_layout.setContentsMargins(0, 0, 0, 0)
        btn_limpiar_layout.setSpacing(4)
        limpiar_spacer = StyledLabel("", size=WindowsPhoneTheme.FONT_SIZE_SMALL)
        btn_limpiar_layout.addWidget(limpiar_spacer)
        btn_limpiar = QPushButton("Limpiar")
        btn_limpiar.setMinimumHeight(40)
        btn_limpiar.setMinimumWidth(100)
        btn_limpiar.setObjectName("tileButton")
        btn_limpiar.setProperty("tileColor", WindowsPhoneTheme.TILE_ORANGE)
        btn_limpiar.clicked.connect(self.limpiar_filtros)
        btn_limpiar_layout.addWidget(btn_limpiar)
        filters_layout.addWidget(btn_limpiar_container)
        
        parent_layout.addWidget(filters_panel)
        
    def create_history_table(self, parent_layout):
        """Crear tabla de historial"""
        self.history_table = QTableWidget()
        self.history_table.setColumnCount(6)
        self.history_table.setHorizontalHeaderLabels([
            "ID", "Fecha", "Hora", "Total", "Usuario", "Detalles"
        ])
        
        # Configurar para que no sea editable
        self.history_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.history_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.history_table.setSelectionMode(QAbstractItemView.SingleSelection)
        
        header = self.history_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.Stretch)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        
        parent_layout.addWidget(self.history_table)
    
    def on_search_changed(self):
        """Reiniciar timer cuando cambia el texto de búsqueda"""
        self.scanner_timer.start()
    
    def limpiar_filtros(self):
        """Limpiar todos los filtros"""
        self.search_bar.clear()
        self.usuario_combo.setCurrentIndex(0)
        self.fecha_desde.setDate(QDate.currentDate().addDays(-30))
        self.fecha_hasta.setDate(QDate.currentDate())
        self.cargar_historial_completo()
    
    def cargar_historial_completo(self):
        """Cargar historial completo de ventas desde la base de datos"""
        try:
            fecha_desde = self.fecha_desde.date().toPython()
            fecha_hasta = self.fecha_hasta.date().toPython()
            
            response = self.pg_manager.client.table('ventas').select(
                'id_venta, fecha, total, usuarios(nombre_completo)'
            ).gte('fecha', f'{fecha_desde}T00:00:00').lte(
                'fecha', f'{fecha_hasta}T23:59:59'
            ).order('fecha', desc=True).execute()
            
            self.ventas_data = response.data or []
            
            # Cargar usuarios para el filtro
            self.cargar_usuarios_filtro()
            
            # Aplicar filtros
            self.aplicar_filtros()
            
        except Exception as e:
            logging.error(f"Error cargando historial: {e}")
            show_warning_dialog(self, "Error", f"Error al cargar historial: {e}")
    
    def cargar_usuarios_filtro(self):
        """Cargar lista de usuarios para el filtro"""
        try:
            # Obtener usuarios únicos de las ventas
            usuarios = set()
            for venta in self.ventas_data:
                if venta.get('usuarios') and isinstance(venta['usuarios'], dict):
                    nombre = venta['usuarios'].get('nombre_completo')
                    if nombre:
                        usuarios.add(nombre)
            
            # Actualizar combo
            self.usuario_combo.clear()
            self.usuario_combo.addItem("Todos")
            for usuario in sorted(usuarios):
                self.usuario_combo.addItem(usuario)
                
        except Exception as e:
            logging.error(f"Error cargando usuarios para filtro: {e}")
    
    def aplicar_filtros(self):
        """Aplicar filtros a los datos de ventas"""
        try:
            # Obtener texto de búsqueda
            search_text = self.search_bar.text().lower().strip()
            usuario_filtro = self.usuario_combo.currentText()
            
            # Filtrar datos
            ventas_filtradas = self.ventas_data
            
            # Filtro por usuario
            if usuario_filtro and usuario_filtro != "Todos":
                ventas_filtradas = [
                    v for v in ventas_filtradas 
                    if v.get('usuarios', {}).get('nombre_completo', '') == usuario_filtro
                ]
            
            # Filtro por búsqueda de texto
            if search_text:
                ventas_filtradas = [
                    v for v in ventas_filtradas
                    if (
                        search_text in str(v.get('id_venta', '')).lower() or
                        search_text in str(v.get('total', '')).lower() or
                        search_text in v.get('usuarios', {}).get('nombre_completo', '').lower()
                    )
                ]
            
            # Actualizar tabla
            self.actualizar_tabla(ventas_filtradas)
            
            # Actualizar info
            total_ventas = sum(v.get('total', 0) for v in ventas_filtradas)
            self.info_label.setText(
                f"Mostrando {len(ventas_filtradas)} de {len(self.ventas_data)} ventas  |  "
                f"Total: ${total_ventas:,.2f}"
            )
            
        except Exception as e:
            logging.error(f"Error aplicando filtros: {e}")
    
    def actualizar_tabla(self, ventas):
        """Actualizar tabla con las ventas filtradas"""
        try:
            self.history_table.setRowCount(len(ventas))
            
            for row, venta in enumerate(ventas):
                self.history_table.setRowHeight(row, 55)
                
                # ID
                self.history_table.setItem(row, 0, QTableWidgetItem(str(venta['id_venta'])))
                
                # Fecha
                fecha = venta['fecha'].strftime("%d/%m/%Y") if isinstance(venta['fecha'], datetime) else str(venta['fecha'])
                self.history_table.setItem(row, 1, QTableWidgetItem(fecha))
                
                # Hora
                hora = venta['fecha'].strftime("%H:%M") if isinstance(venta['fecha'], datetime) else "N/A"
                self.history_table.setItem(row, 2, QTableWidgetItem(hora))
                
                # Total
                total_item = QTableWidgetItem(f"${venta['total']:.2f}")
                total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.history_table.setItem(row, 3, total_item)
                
                # Usuario
                usuario_name = venta.get('usuarios', {}).get('nombre_completo', 'N/A') if isinstance(venta.get('usuarios'), dict) else venta.get('usuario', 'N/A')
                self.history_table.setItem(row, 4, QTableWidgetItem(usuario_name))
                
                # Botón detalles con icono
                btn_detalles = QPushButton()
                btn_detalles.setIcon(qta.icon('fa5s.eye', color='white'))
                btn_detalles.setToolTip("Ver detalle de la venta")
                btn_detalles.setFixedWidth(40)
                btn_detalles.setMinimumHeight(35)
                btn_detalles.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {WindowsPhoneTheme.TILE_BLUE};
                        color: white;
                        border: none;
                        border-radius: 3px;
                    }}
                    QPushButton:hover {{
                        background-color: #1976d2;
                    }}
                """)
                btn_detalles.clicked.connect(lambda checked, vid=venta['id_venta']: self.ver_detalles(vid))
                self.history_table.setCellWidget(row, 5, btn_detalles)
                
        except Exception as e:
            logging.error(f"Error cargando historial de ventas: {e}")
            show_warning_dialog(self, "Error", f"No se pudo cargar el historial: {e}")
            
    def ver_detalles(self, venta_id):
        """Ver detalles de una venta"""
        try:
            response = self.pg_manager.client.table('detalles_venta').select(
                'codigo_interno, tipo_producto, cantidad, precio_unitario, subtotal_linea, nombre_producto, descripcion_producto'
            ).eq('id_venta', venta_id).execute()
            
            detalles = response.data or []
            
            if not detalles:
                show_info_dialog(self, "Detalles", f"No se encontraron detalles para la venta ID: {venta_id}")
                return
                
            # Crear mensaje con los detalles
            mensaje = f"Detalles de la venta ID: {venta_id}\n\n"
            mensaje += "Producto\t\tCantidad\tPrecio\tSubtotal\n"
            mensaje += "------------------------------------------------\n"
            
            for detalle in detalles:
                nombre = detalle['nombre_producto'][:20] + "..." if len(detalle['nombre_producto']) > 20 else detalle['nombre_producto']
                mensaje += f"{nombre}\t{detalle['cantidad']}\t${detalle['precio_unitario']:.2f}\t${detalle['subtotal_linea']:.2f}\n"
            
            show_info_dialog(self, "Detalles de Venta", mensaje)
            
        except Exception as e:
            logging.error(f"Error obteniendo detalles de venta: {e}")
            show_warning_dialog(self, "Error", f"No se pudieron obtener los detalles: {e}")
        
    def exportar_datos(self):
        """Exportar datos a archivo"""
        try:
            from datetime import datetime
            import os
            
            # Verificar si openpyxl está disponible
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                show_warning_dialog(
                    self,
                    "Biblioteca requerida",
                    "Para exportar a Excel necesitas instalar openpyxl",
                    detail="Ejecuta: pip install openpyxl"
                )
                return
            
            # Obtener rango de fechas
            fecha_desde = self.fecha_desde.date().toPython()
            fecha_hasta = self.fecha_hasta.date().toPython()
            
            # Obtener datos de ventas desde Supabase
            response = self.pg_manager.client.table('ventas').select(
                'id_venta, fecha, total, usuarios(nombre_completo)'
            ).gte('fecha', f'{fecha_desde}T00:00:00').lte(
                'fecha', f'{fecha_hasta}T23:59:59'
            ).order('fecha', desc=True).execute()
            
            ventas = response.data or []
            
            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Historial de Ventas"
            
            # Estilos
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezados
            headers = ["ID Venta", "Fecha", "Hora", "Total", "Usuario"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Datos
            for row, venta in enumerate(ventas, 2):
                ws.cell(row=row, column=1, value=venta['id_venta']).border = border
                
                # Fecha
                fecha = venta['fecha'].strftime("%d/%m/%Y") if isinstance(venta['fecha'], datetime) else str(venta['fecha'])
                ws.cell(row=row, column=2, value=fecha).border = border
                
                # Hora
                hora = venta['fecha'].strftime("%H:%M") if isinstance(venta['fecha'], datetime) else "N/A"
                ws.cell(row=row, column=3, value=hora).border = border
                
                # Total
                total_cell = ws.cell(row=row, column=4, value=venta['total'])
                total_cell.number_format = '$#,##0.00'
                total_cell.border = border
                
                # Usuario
                ws.cell(row=row, column=5, value=venta.get('usuario', 'N/A')).border = border
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 25
            
            # Guardar archivo
            fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            filename = os.path.join(desktop, f"historial_ventas_{fecha_str}.xlsx")
            
            wb.save(filename)
            
            show_info_dialog(
                self,
                "Exportación completada",
                f"El historial de ventas ha sido exportado exitosamente",
                detail=f"Archivo guardado en:\n{filename}"
            )
            
            logging.info(f"Historial de ventas exportado: {filename}")
            
        except Exception as e:
            logging.error(f"Error exportando datos: {e}")
            show_warning_dialog(
                self,
                "Error al exportar",
                "No se pudo exportar el historial de ventas",
                detail=str(e)
            )