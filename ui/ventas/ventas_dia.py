"""
Ventana de Ventas del Día para HTF POS
Usando componentes reutilizables del sistema de diseño
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, 
    QTableWidget, QTableWidgetItem,
    QHeaderView, QSizePolicy, QDialog, QPushButton, QLabel
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QFont
import logging
from datetime import datetime, date
import qtawesome as qta

# Importar componentes del sistema de diseño
from ui.components import (
    WindowsPhoneTheme,
    TileButton,
    InfoTile,
    create_page_layout,
    show_info_dialog,
    show_warning_dialog,
    SectionTitle,
    ContentPanel
)


class VentasDiaWindow(QWidget):
    """Widget para ver ventas del turno"""
    
    cerrar_solicitado = Signal()
    
    def __init__(self, pg_manager, supabase_service, user_data, turno_id=None, parent=None):
        super().__init__(parent)
        self.pg_manager = pg_manager
        self.supabase_service = supabase_service
        self.user_data = user_data
        self.turno_id = turno_id  # ID del turno actual
        
        # Configurar política de tamaño
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        self.setup_ui()
        
    def setup_ui(self):
        """Configurar interfaz de ventas del turno"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Contenido
        content = QWidget()
        content_layout = create_page_layout("VENTAS DEL TURNO ACTUAL")
        content.setLayout(content_layout)
        
        # Widgets de resumen (sin promedio)
        self.create_summary_widgets(content_layout)
        
        # Botones de acción
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(WindowsPhoneTheme.TILE_SPACING)
        
        btn_detalle = TileButton("Ver Detalle", "fa5s.list", WindowsPhoneTheme.TILE_BLUE)
        btn_detalle.clicked.connect(self.ver_detalle_ventas)
        
        btn_imprimir = TileButton("Imprimir", "fa5s.print", WindowsPhoneTheme.TILE_GREEN)
        btn_imprimir.clicked.connect(self.imprimir_reporte)
        
        btn_cerrar = TileButton("Cerrar", "fa5s.times", WindowsPhoneTheme.TILE_RED)
        btn_cerrar.clicked.connect(self.cerrar_solicitado.emit)
        
        buttons_layout.addWidget(btn_detalle)
        buttons_layout.addWidget(btn_imprimir)
        buttons_layout.addWidget(btn_cerrar)
        
        content_layout.addLayout(buttons_layout)
        layout.addWidget(content)
        
        # Cargar datos iniciales
        self.actualizar_datos()
        
    def create_summary_widgets(self, parent_layout):
        """Crear widgets de resumen"""
        widgets_layout = QHBoxLayout()
        widgets_layout.setSpacing(WindowsPhoneTheme.TILE_SPACING)
        
        # Total vendido
        self.total_tile = InfoTile("TOTAL VENDIDO", "fa5s.dollar-sign", WindowsPhoneTheme.TILE_GREEN)
        self.total_value = self.total_tile.add_main_value("$0.00")
        widgets_layout.addWidget(self.total_tile)
        
        # Número de ventas
        self.ventas_tile = InfoTile("COMANDAS", "fa5s.shopping-cart", WindowsPhoneTheme.TILE_BLUE)
        self.ventas_count = self.ventas_tile.add_main_value("0")
        self.ventas_tile.add_secondary_value("comandas del turno")
        widgets_layout.addWidget(self.ventas_tile)
        
        parent_layout.addLayout(widgets_layout)
        
    def actualizar_datos(self):
        """Actualizar datos de ventas del turno"""
        try:
            # Verificar que haya un turno activo
            if not self.turno_id:
                show_warning_dialog(
                    self,
                    "Turno No Disponible",
                    "No hay un turno de caja abierto.",
                    "Debes abrir un turno antes de ver las ventas."
                )
                self.total_value.setText("$0.00")
                self.ventas_count.setText("0")
                return
            
            # Obtener ventas del turno usando Supabase
            response = self.pg_manager.client.table('ventas').select(
                'id_venta, fecha, total, usuarios(nombre_completo)'
            ).eq('id_turno', self.turno_id).order('fecha', desc=True).execute()
            
            ventas = response.data or []
            total_vendido = sum(v.get('total', 0) for v in ventas)
            num_ventas = len(ventas)
            
            self.total_value.setText(f"${total_vendido:.2f}")
            self.ventas_count.setText(str(num_ventas))

        except Exception as e:
            logging.error(f"Error actualizando datos: {e}")
            show_warning_dialog(self, "Error", f"Error al cargar datos: {e}")

    def ver_detalle_ventas(self):
        """Abrir ventana de detalle de ventas del turno"""
        if not self.turno_id:
            show_warning_dialog(
                self,
                "Turno No Disponible",
                "No hay un turno de caja abierto."
            )
            return
        
        dialog = DetalleVentasDiaDialog(self.pg_manager, self.turno_id, self)
        dialog.exec()
            
    def imprimir_reporte(self):
        """Imprimir reporte de ventas del turno"""
        try:
            if not self.turno_id:
                show_warning_dialog(
                    self,
                    "Turno No Disponible",
                    "No hay un turno de caja abierto."
                )
                return
            
            from datetime import datetime
            import os
            
            # Verificar si openpyxl está disponible
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except ImportError:
                show_warning_dialog(
                    self,
                    "Biblioteca requerida",
                    "Para exportar a Excel necesitas instalar openpyxl",
                    detail="Ejecuta: pip install openpyxl"
                )
                return
            
            # Obtener ventas del turno desde Supabase
            response = self.pg_manager.client.table('ventas').select(
                'id_venta, fecha, total, usuarios(nombre_completo)'
            ).eq('id_turno', self.turno_id).order('fecha', desc=True).execute()
            
            ventas = response.data or []
            
            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Ventas del Día"
            
            # Estilos
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Encabezados
            headers = ["ID Venta", "Fecha", "Hora", "Total", "Usuario"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border
            
            # Datos
            for row, venta in enumerate(ventas, 2):
                ws.cell(row=row, column=1, value=venta['id_venta']).border = border
                
                # Fecha
                fecha = venta['fecha'].strftime("%d/%m/%Y") if isinstance(venta['fecha'], datetime) else str(venta['fecha'])
                ws.cell(row=row, column=2, value=fecha).border = border
                
                # Hora
                hora = venta['fecha'].strftime("%H:%M") if isinstance(venta['fecha'], datetime) else "N/A"
                ws.cell(row=row, column=3, value=hora).border = border
                
                # Total
                total_cell = ws.cell(row=row, column=4, value=venta['total'])
                total_cell.number_format = '$#,##0.00'
                total_cell.border = border
                
                # Usuario
                ws.cell(row=row, column=5, value=venta.get('usuario', 'N/A')).border = border
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 25
            
            # Guardar archivo
            fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            filename = os.path.join(desktop, f"ventas_turno_{self.turno_id}_{fecha_str}.xlsx")
            
            wb.save(filename)
            
            show_info_dialog(
                self,
                "Exportación completada",
                f"El reporte de ventas del turno ha sido exportado exitosamente",
                detail=f"Archivo guardado en:\n{filename}"
            )
            
            logging.info(f"Ventas del turno {self.turno_id} exportadas: {filename}")
            
        except Exception as e:
            logging.error(f"Error exportando datos: {e}")
            show_warning_dialog(
                self,
                "Error al exportar",
                "No se pudo exportar el reporte de ventas",
                detail=str(e)
            )


class DetalleVentasDiaDialog(QDialog):
    """Diálogo para ver el detalle de todas las ventas del turno"""
    
    def __init__(self, pg_manager, turno_id, parent=None):
        super().__init__(parent)
        self.pg_manager = pg_manager
        self.turno_id = turno_id
        
        self.setWindowTitle("Detalle de Ventas del Turno")
        self.setMinimumSize(900, 600)
        
        self.setup_ui()
        self.cargar_ventas()
        
    def setup_ui(self):
        """Configurar interfaz del diálogo"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(WindowsPhoneTheme.MARGIN_MEDIUM,
                                 WindowsPhoneTheme.MARGIN_MEDIUM,
                                 WindowsPhoneTheme.MARGIN_MEDIUM,
                                 WindowsPhoneTheme.MARGIN_MEDIUM)
        layout.setSpacing(WindowsPhoneTheme.MARGIN_SMALL)
        
        # Título
        title = SectionTitle("COMANDAS DEL TURNO")
        layout.addWidget(title)
        
        # Tabla de comandas
        self.tabla_ventas = QTableWidget()
        self.tabla_ventas.setColumnCount(6)
        self.tabla_ventas.setHorizontalHeaderLabels([
            "ID", "Hora", "Total", "Usuario", "Estado", "Acción"
        ])
        
        header = self.tabla_ventas.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.Fixed)
        header.resizeSection(5, 120)
        
        layout.addWidget(self.tabla_ventas)
        
        # Botón cerrar
        btn_cerrar = TileButton("Cerrar", "fa5s.times", WindowsPhoneTheme.TILE_RED)
        btn_cerrar.setMaximumWidth(200)
        btn_cerrar.clicked.connect(self.accept)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(btn_cerrar)
        layout.addLayout(btn_layout)
        
    def cargar_ventas(self):
        """Cargar ventas del turno"""
        try:
            response = self.pg_manager.client.table('ventas').select(
                'id_venta, fecha, total, usuarios(nombre_completo)'
            ).eq('id_turno', self.turno_id).order('fecha', desc=True).execute()
            
            ventas = response.data or []
            
            self.tabla_ventas.setRowCount(len(ventas))
            
            for row, venta in enumerate(ventas):
                # Establecer altura de fila
                self.tabla_ventas.setRowHeight(row, 55)
                
                # ID
                self.tabla_ventas.setItem(row, 0, QTableWidgetItem(str(venta['id_venta'])))
                
                # Hora
                hora = venta['fecha'].strftime("%H:%M") if isinstance(venta['fecha'], datetime) else str(venta['fecha'])
                self.tabla_ventas.setItem(row, 1, QTableWidgetItem(hora))
                
                # Total
                total_item = QTableWidgetItem(f"${venta['total']:.2f}")
                total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tabla_ventas.setItem(row, 2, total_item)
                
                # Usuario
                usuario_name = venta.get('usuarios', {}).get('nombre_completo', 'N/A') if isinstance(venta.get('usuarios'), dict) else venta.get('usuario', 'N/A')
                self.tabla_ventas.setItem(row, 3, QTableWidgetItem(usuario_name))
                
                # Estado
                self.tabla_ventas.setItem(row, 4, QTableWidgetItem("Completada"))
                
                # Botón ver detalle con icono
                btn_ver = QPushButton()
                btn_ver.setIcon(qta.icon('fa5s.eye', color='white'))
                btn_ver.setToolTip("Ver detalle de la venta")
                btn_ver.setFixedWidth(40)
                btn_ver.setMinimumHeight(35)
                btn_ver.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {WindowsPhoneTheme.TILE_BLUE};
                        color: white;
                        border: none;
                        border-radius: 3px;
                    }}
                    QPushButton:hover {{
                        background-color: #1976d2;
                    }}
                """)
                btn_ver.clicked.connect(lambda checked, v_id=venta['id_venta']: self.ver_detalle_comanda(v_id))
                self.tabla_ventas.setCellWidget(row, 5, btn_ver)
                
        except Exception as e:
            logging.error(f"Error cargando ventas: {e}")
            show_warning_dialog(self, "Error", f"Error al cargar ventas: {e}")
            
    def ver_detalle_comanda(self, venta_id):
        """Abrir ventana con detalle de una comanda específica"""
        dialog = DetalleComandaDialog(self.pg_manager, venta_id, self)
        dialog.exec()


class DetalleComandaDialog(QDialog):
    """Diálogo para ver el detalle de una comanda individual"""
    
    def __init__(self, pg_manager, venta_id, parent=None):
        super().__init__(parent)
        self.pg_manager = pg_manager
        self.venta_id = venta_id
        
        self.setWindowTitle(f"Detalle de Comanda #{venta_id}")
        self.setMinimumSize(700, 500)
        
        self.setup_ui()
        self.cargar_detalle()
        
    def setup_ui(self):
        """Configurar interfaz del diálogo"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(WindowsPhoneTheme.MARGIN_MEDIUM,
                                 WindowsPhoneTheme.MARGIN_MEDIUM,
                                 WindowsPhoneTheme.MARGIN_MEDIUM,
                                 WindowsPhoneTheme.MARGIN_MEDIUM)
        layout.setSpacing(WindowsPhoneTheme.MARGIN_SMALL)
        
        # Título
        self.title_label = SectionTitle(f"COMANDA #{self.venta_id}")
        layout.addWidget(self.title_label)
        
        # Panel de información de cabecera
        info_panel = ContentPanel()
        info_layout = QVBoxLayout(info_panel)
        info_layout.setContentsMargins(20, 20, 20, 20)
        info_layout.setSpacing(10)
        
        self.fecha_label = QLabel("Fecha: -")
        self.fecha_label.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        
        self.usuario_label = QLabel("Usuario: -")
        self.usuario_label.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_NORMAL))
        
        self.total_label = QLabel("Total: $0.00")
        self.total_label.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_LARGE, QFont.Bold))
        
        info_layout.addWidget(self.fecha_label)
        info_layout.addWidget(self.usuario_label)
        info_layout.addWidget(self.total_label)
        
        layout.addWidget(info_panel)
        
        # Tabla de productos
        productos_title = QLabel("PRODUCTOS")
        productos_title.setFont(QFont(WindowsPhoneTheme.FONT_FAMILY, WindowsPhoneTheme.FONT_SIZE_SUBTITLE, QFont.Bold))
        layout.addWidget(productos_title)
        
        self.tabla_productos = QTableWidget()
        self.tabla_productos.setColumnCount(4)
        self.tabla_productos.setHorizontalHeaderLabels([
            "Producto", "Cantidad", "Precio Unit.", "Subtotal"
        ])
        
        header = self.tabla_productos.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        
        layout.addWidget(self.tabla_productos)
        
        # Botón cerrar
        btn_cerrar = TileButton("Cerrar", "fa5s.times", WindowsPhoneTheme.TILE_RED)
        btn_cerrar.setMaximumWidth(200)
        btn_cerrar.clicked.connect(self.accept)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(btn_cerrar)
        layout.addLayout(btn_layout)
        
    def cargar_detalle(self):
        """Cargar detalle de la comanda"""
        try:
            # Obtener información de la venta
            response = self.pg_manager.client.table('ventas').select(
                'id_venta, fecha, total, usuarios(nombre_completo)'
            ).eq('id_venta', self.venta_id).single().execute()
            
            venta = response.data
            
            if not venta:
                show_warning_dialog(self, "Error", f"No se encontró la venta #{self.venta_id}")
                return
            
            # Actualizar información de cabecera
            fecha_str = venta['fecha'][:16].replace('T', ' ') if isinstance(venta['fecha'], str) else venta['fecha'].strftime("%d/%m/%Y %H:%M")
            usuario_name = venta.get('usuarios', {}).get('nombre_completo', 'N/A') if isinstance(venta.get('usuarios'), dict) else 'N/A'
            self.fecha_label.setText(f"Fecha: {fecha_str}")
            self.usuario_label.setText(f"Usuario: {usuario_name}")
            self.total_label.setText(f"Total: ${venta['total']:.2f}")
            
            # Obtener items de la venta
            response_items = self.pg_manager.client.table('detalles_venta').select(
                'codigo_interno, tipo_producto, cantidad, precio_unitario, subtotal_linea, nombre_producto, descripcion_producto'
            ).eq('id_venta', self.venta_id).execute()
            
            items = response_items.data or []
            
            self.tabla_productos.setRowCount(len(items))
            
            for row, item in enumerate(items):
                # Producto
                self.tabla_productos.setItem(row, 0, QTableWidgetItem(item['nombre_producto']))
                
                # Cantidad
                cantidad_item = QTableWidgetItem(str(item['cantidad']))
                cantidad_item.setTextAlignment(Qt.AlignCenter)
                self.tabla_productos.setItem(row, 1, cantidad_item)
                
                # Precio unitario
                precio_item = QTableWidgetItem(f"${item['precio_unitario']:.2f}")
                precio_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tabla_productos.setItem(row, 2, precio_item)
                
                # Subtotal
                subtotal_item = QTableWidgetItem(f"${item['subtotal_linea']:.2f}")
                subtotal_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.tabla_productos.setItem(row, 3, subtotal_item)
                
        except Exception as e:
            logging.error(f"Error cargando detalle: {e}")
            show_warning_dialog(self, "Error", f"Error al cargar detalle: {e}")