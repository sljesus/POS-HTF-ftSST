"""
Módulo para gestionar importación y exportación de Excel
Funciona con openpyxl y pandas
"""

import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelManager:
    """Gestor de archivos Excel para importar y exportar datos"""
    
    def __init__(self, ruta_base: str = None):
        """
        Inicializar el gestor
        
        Args:
            ruta_base: Ruta base para guardar archivos (default: carpeta actual)
        """
        self.ruta_base = ruta_base or os.getcwd()
        self.asegurar_carpeta()
    
    def asegurar_carpeta(self):
        """Crear carpeta de exportación si no existe"""
        carpeta = os.path.join(self.ruta_base, "exportes_excel")
        Path(carpeta).mkdir(exist_ok=True)
        self.carpeta_exportes = carpeta
    
    # ===== EXPORTAR =====
    
    def exportar_simple(self, datos: List[Dict], nombre_archivo: str, 
                       titulo_hoja: str = "Datos") -> str:
        """
        Exportar datos simples a Excel
        
        Args:
            datos: Lista de diccionarios
            nombre_archivo: Nombre del archivo (sin extensión)
            titulo_hoja: Nombre de la hoja
            
        Returns:
            Ruta del archivo creado
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")
        
        if not datos:
            logger.warning("No hay datos para exportar")
            return None
        
        try:
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = titulo_hoja
            
            # Obtener encabezados
            encabezados = list(datos[0].keys())
            
            # Escribir encabezados con formato
            for col_idx, encabezado in enumerate(encabezados, 1):
                celda = ws.cell(row=1, column=col_idx)
                celda.value = encabezado
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                celda.alignment = Alignment(horizontal="center", vertical="center")
            
            # Escribir datos
            for row_idx, fila in enumerate(datos, 2):
                for col_idx, encabezado in enumerate(encabezados, 1):
                    valor = fila.get(encabezado, "")
                    celda = ws.cell(row=row_idx, column=col_idx)
                    celda.value = valor
                    celda.alignment = Alignment(horizontal="left", vertical="center")
            
            # Ajustar ancho de columnas
            for col_idx in range(1, len(encabezados) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
            
            # Guardar
            ruta_archivo = os.path.join(self.carpeta_exportes, f"{nombre_archivo}.xlsx")
            wb.save(ruta_archivo)
            logger.info(f"✅ Excel exportado: {ruta_archivo}")
            return ruta_archivo
            
        except Exception as e:
            logger.error(f"❌ Error al exportar: {e}")
            raise
    
    def exportar_con_formato(self, datos: List[Dict], nombre_archivo: str,
                           titulo: str = "", subtitulo: str = "") -> str:
        """
        Exportar con formato profesional
        
        Args:
            datos: Lista de diccionarios
            nombre_archivo: Nombre del archivo
            titulo: Título del reporte
            subtitulo: Subtítulo
            
        Returns:
            Ruta del archivo
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado")
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte"
            
            # Título
            if titulo:
                ws.merge_cells('A1:H1')
                celda_titulo = ws['A1']
                celda_titulo.value = titulo
                celda_titulo.font = Font(size=14, bold=True, color="FFFFFF")
                celda_titulo.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
                celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 25
            
            # Subtítulo y fecha
            fila_actual = 2
            if subtitulo:
                ws.merge_cells(f'A{fila_actual}:H{fila_actual}')
                celda_sub = ws[f'A{fila_actual}']
                celda_sub.value = subtitulo
                celda_sub.font = Font(size=10, italic=True)
                celda_sub.alignment = Alignment(horizontal="center")
                fila_actual += 1
            
            # Fecha de generación
            ws.merge_cells(f'A{fila_actual}:H{fila_actual}')
            celda_fecha = ws[f'A{fila_actual}']
            celda_fecha.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            celda_fecha.font = Font(size=9, italic=True, color="666666")
            celda_fecha.alignment = Alignment(horizontal="right")
            fila_actual += 2
            
            # Encabezados
            if datos:
                encabezados = list(datos[0].keys())
                for col_idx, encabezado in enumerate(encabezados, 1):
                    celda = ws.cell(row=fila_actual, column=col_idx)
                    celda.value = encabezado
                    celda.font = Font(bold=True, color="FFFFFF", size=11)
                    celda.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    celda.alignment = Alignment(horizontal="center", vertical="center")
                    celda.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Datos
                for row_offset, fila_datos in enumerate(datos, 1):
                    for col_idx, encabezado in enumerate(encabezados, 1):
                        valor = fila_datos.get(encabezado, "")
                        celda = ws.cell(row=fila_actual + row_offset, column=col_idx)
                        celda.value = valor
                        celda.alignment = Alignment(horizontal="left", vertical="center")
                        celda.border = Border(
                            left=Side(style='thin', color="CCCCCC"),
                            right=Side(style='thin', color="CCCCCC"),
                            bottom=Side(style='thin', color="CCCCCC")
                        )
                
                # Ajustar ancho
                for col_idx in range(1, len(encabezados) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
            
            ruta_archivo = os.path.join(self.carpeta_exportes, f"{nombre_archivo}.xlsx")
            wb.save(ruta_archivo)
            logger.info(f"✅ Reporte exportado: {ruta_archivo}")
            return ruta_archivo
            
        except Exception as e:
            logger.error(f"❌ Error al exportar con formato: {e}")
            raise
    
    # ===== IMPORTAR =====
    
    def importar_simple(self, ruta_archivo: str) -> List[Dict]:
        """
        Importar datos desde Excel
        
        Args:
            ruta_archivo: Ruta del archivo Excel
            
        Returns:
            Lista de diccionarios con los datos
        """
        if not PANDAS_AVAILABLE:
            logger.warning("pandas no disponible, usando openpyxl")
            return self._importar_con_openpyxl(ruta_archivo)
        
        try:
            df = pd.read_excel(ruta_archivo)
            datos = df.to_dict(orient='records')
            logger.info(f"✅ {len(datos)} registros importados")
            return datos
        except Exception as e:
            logger.error(f"❌ Error al importar: {e}")
            raise
    
    def _importar_con_openpyxl(self, ruta_archivo: str) -> List[Dict]:
        """Importar sin pandas (openpyxl)"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("Se requiere openpyxl o pandas")
        
        try:
            wb = openpyxl.load_workbook(ruta_archivo)
            ws = wb.active
            
            # Obtener encabezados
            encabezados = []
            for celda in ws[1]:
                encabezados.append(celda.value)
            
            # Obtener datos
            datos = []
            for fila in ws.iter_rows(min_row=2, values_only=False):
                fila_dict = {}
                for col_idx, celda in enumerate(fila):
                    if col_idx < len(encabezados):
                        fila_dict[encabezados[col_idx]] = celda.value
                datos.append(fila_dict)
            
            logger.info(f"✅ {len(datos)} registros importados")
            return datos
            
        except Exception as e:
            logger.error(f"❌ Error al importar: {e}")
            raise
    
    def importar_con_validacion(self, ruta_archivo: str, 
                               columnas_requeridas: List[str]) -> tuple[bool, List[Dict], str]:
        """
        Importar con validación de columnas
        
        Args:
            ruta_archivo: Ruta del archivo
            columnas_requeridas: Columnas que debe tener
            
        Returns:
            (éxito, datos, mensaje)
        """
        try:
            datos = self.importar_simple(ruta_archivo)
            
            if not datos:
                return False, [], "El archivo está vacío"
            
            # Verificar columnas
            columnas_presentes = set(datos[0].keys())
            columnas_faltantes = set(columnas_requeridas) - columnas_presentes
            
            if columnas_faltantes:
                msg = f"Columnas faltantes: {', '.join(columnas_faltantes)}"
                logger.error(f"❌ {msg}")
                return False, [], msg
            
            logger.info(f"✅ Validación exitosa: {len(datos)} registros válidos")
            return True, datos, "OK"
            
        except Exception as e:
            logger.error(f"❌ Error en validación: {e}")
            return False, [], str(e)


# ===== EJEMPLOS DE USO =====

if __name__ == "__main__":
    # Crear gestor
    gestor = ExcelManager()
    
    # Ejemplo 1: Exportar simple
    datos_ejemplo = [
        {"Nombre": "Juan", "Apellido": "Pérez", "Membresía": "Gold", "Cantidad": 15},
        {"Nombre": "María", "Apellido": "García", "Membresía": "Silver", "Cantidad": 8},
        {"Nombre": "Carlos", "Apellido": "López", "Membresía": "Bronze", "Cantidad": 3},
    ]
    
    ruta = gestor.exportar_simple(datos_ejemplo, "miembros_simple", "Miembros")
    print(f"Archivo guardado: {ruta}\n")
    
    # Ejemplo 2: Exportar con formato
    ruta = gestor.exportar_con_formato(
        datos_ejemplo,
        "reporte_miembros",
        titulo="REPORTE DE MIEMBROS",
        subtitulo="Gimnasio HTF - Diciembre 2025"
    )
    print(f"Reporte guardado: {ruta}\n")
    
    # Ejemplo 3: Importar
    # datos_importados = gestor.importar_simple(ruta)
    # print("Datos importados:", datos_importados)
